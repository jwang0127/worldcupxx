#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from __future__ import annotations

import html
import json
import os
import re
import shutil
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SCHEDULE_MD = ROOT / "2026世界杯淘汰赛赛程表_中文版_Codex.md"
SCHEDULE_XLSX = ROOT / "2026世界杯淘汰赛赛程表_中文版_Codex.xlsx"
MODEL_MD = ROOT / "world_cup_knockout_model_optimization_codex.md"
GROUP_MODEL_MD = ROOT / "SHAREABLE_PREDICTION_REVIEW_LOGIC.md"
SHARE_NOTE_MD = ROOT / "1.md"
DATA_DIR = ROOT / "data"
PREDICTION_DATE = "20260629"
FIRST_GAME_ID = "53452545"
PREDICTION_GAME_IDS = ["53452545", "53452557", "53452541", "53452547", "53452561", "53452543", "53452563"]
GLOBAL_SCHEDULE: list[dict[str, Any]] = []
PARLAY_POLICY = {
    "min_enabled_match_count": 2,
    "single_match_mode": "regular_prediction_only",
    "required_markets": ["crs", "ttg", "hafu"],
    "odds_source": "Sporttery getMatchCalculatorV1 website API via scripts/fetch_sporttery.ps1",
    "parlay_types": ["比分三串一", "总进球数三串一", "半场胜平负三串一"],
}


TEAM_ALIAS = {
    "刚果民主共和国": "刚果(金)",
    "DR Congo": "刚果(金)",
    "Congo DR": "刚果(金)",
    "United States": "美国",
    "South Africa": "南非",
    "Canada": "加拿大",
}


def esc(value: Any) -> str:
    return html.escape("" if value is None else str(value), quote=True)


def read_text(path: Path) -> str:
    try:
        return path.read_text(encoding="utf-8-sig")
    except UnicodeDecodeError:
        return path.read_text(encoding="utf-8")


def read_json(path: Path) -> Any:
    return json.loads(read_text(path))


def odds_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def fmt_odds(value: float | None) -> str:
    return "-" if value is None else f"{value:.2f}"


def score_total(score: str) -> int | None:
    matched = re.match(r"^\s*(\d+)\s*[-:：]\s*(\d+)\s*$", str(score))
    if not matched:
        return None
    return int(matched.group(1)) + int(matched.group(2))


def half_direction_from_text(value: str) -> str:
    matched = re.search(r"(\d+)\s*[-:：]\s*(\d+)", str(value))
    if not matched:
        return "draw"
    home_goals = int(matched.group(1))
    away_goals = int(matched.group(2))
    if home_goals > away_goals:
        return "home"
    if home_goals < away_goals:
        return "away"
    return "draw"


def half_direction_label(direction: str) -> str:
    return {"home": "半场主胜", "draw": "半场平", "away": "半场客胜"}.get(direction, direction)


def outcome_short_label(direction: str) -> str:
    return {"home": "胜", "draw": "平", "away": "负"}.get(direction, direction)


def load_live_odds(date_label: str) -> dict[str, Any]:
    path = DATA_DIR / f"{date_label}.json"
    if not path.exists():
        return {}
    try:
        payload = read_json(path)
    except Exception:
        return {}
    by_id: dict[str, Any] = {}
    for item in payload.get("matches", []):
        match_id = str(item.get("id", "")).lstrip("0")
        if match_id:
            by_id[match_id] = item
        key = f"{item.get('home', '')}|{item.get('away', '')}"
        by_id[key] = item
    return by_id


def odds_match_for_prediction(item: dict[str, Any], odds_lookup: dict[str, Any]) -> dict[str, Any] | None:
    match_no = str(item.get("match_no", "")).lstrip("0")
    if match_no and match_no in odds_lookup:
        return odds_lookup[match_no]
    key = f"{item.get('home_team', '')}|{item.get('away_team', '')}"
    return odds_lookup.get(key)


def hafu_key(half_direction: str, full_direction: str) -> str:
    prefix = {"home": "h", "draw": "d", "away": "a"}.get(half_direction, "d")
    suffix = {"home": "h", "draw": "d", "away": "a"}.get(full_direction, "d")
    return prefix + suffix


def hafu_label(half_direction: str, full_direction: str) -> str:
    return outcome_short_label(half_direction) + outcome_short_label(full_direction)


def hafu_label_from_key(key: str) -> str:
    if len(str(key)) != 2:
        return str(key)
    reverse = {"h": "home", "d": "draw", "a": "away"}
    return hafu_label(reverse.get(key[0], key[0]), reverse.get(key[1], key[1]))


def hafu_key_from_label_or_key(value: str) -> str:
    text = str(value).strip().lower()
    if re.match(r"^[hda]{2}$", text):
        return text
    mapping = {"胜": "h", "平": "d", "负": "a", "主胜": "h", "客胜": "a"}
    compact = str(value).strip().replace("/", "").replace("-", "").replace(" ", "")
    if len(compact) >= 2:
        left = mapping.get(compact[0])
        right = mapping.get(compact[1])
        if left and right:
            return left + right
    return text


def hafu_exact_odds(hafu: dict[str, Any] | None, half_direction: str, full_direction: str) -> float | None:
    if not hafu:
        return None
    return odds_float(hafu.get(hafu_key(half_direction, full_direction)))


def half_result_label(direction: str) -> str:
    return {"home": "半场主胜", "draw": "半场平", "away": "半场客胜"}.get(direction, direction)


def half_result_implied_odds(hafu: dict[str, Any] | None, half_direction: str) -> float | None:
    if not hafu:
        return None
    prefix = {"home": "h", "draw": "d", "away": "a"}.get(half_direction, "d")
    implied_probability = 0.0
    for suffix in ("h", "d", "a"):
        odd = odds_float(hafu.get(prefix + suffix))
        if odd and odd > 0:
            implied_probability += 1 / odd
    if implied_probability <= 0:
        return None
    return 1 / implied_probability


def product_odds(rows: list[dict[str, Any]]) -> float | None:
    product = 1.0
    for row in rows:
        odd = row.get("odds")
        if odd is None:
            return None
        product *= float(odd)
    return product


def parse_markdown_table(path: Path) -> list[dict[str, str]]:
    lines = read_text(path).splitlines()
    table_lines = [line.strip() for line in lines if line.strip().startswith("|")]
    if len(table_lines) < 3:
        raise RuntimeError(f"Markdown table not found: {path}")

    headers = [cell.strip() for cell in table_lines[0].strip("|").split("|")]
    rows: list[dict[str, str]] = []
    for line in table_lines[2:]:
        if re.match(r"^\|\s*-+", line):
            continue
        cells = [cell.strip() for cell in line.strip("|").split("|")]
        if len(cells) < len(headers):
            cells += [""] * (len(headers) - len(cells))
        item = dict(zip(headers, cells))
        if item.get("比赛ID"):
            rows.append(item)
    return rows


def read_excel_rows(path: Path) -> list[dict[str, str]]:
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [str(ws.cell(1, col).value).strip() for col in range(1, ws.max_column + 1)]
    rows: list[dict[str, str]] = []
    for row_idx in range(2, ws.max_row + 1):
        item = {}
        for col_idx, header in enumerate(headers, start=1):
            value = ws.cell(row_idx, col_idx).value
            item[header] = "" if value is None else str(value).strip()
        if item.get("比赛ID"):
            rows.append(item)
    return rows


def verify_schedule(md_rows: list[dict[str, str]], xlsx_rows: list[dict[str, str]]) -> dict[str, Any]:
    xlsx_by_id = {row["比赛ID"]: row for row in xlsx_rows}
    fields = ["场次编号", "轮次", "名义主队中文", "名义客队中文", "北京时间", "胜者晋级比赛ID"]
    mismatches: list[dict[str, str]] = []
    for row in md_rows:
        other = xlsx_by_id.get(row["比赛ID"])
        if not other:
            mismatches.append({"比赛ID": row["比赛ID"], "字段": "missing-in-xlsx"})
            continue
        for field in fields:
            if str(row.get(field, "")).strip() != str(other.get(field, "")).strip():
                mismatches.append(
                    {
                        "比赛ID": row["比赛ID"],
                        "字段": field,
                        "markdown": row.get(field, ""),
                        "excel": other.get(field, ""),
                    }
                )
    return {"checked": len(md_rows), "mismatches": mismatches}


def normalize_schedule(rows: list[dict[str, str]]) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    for row in rows:
        out.append(
            {
                "game_id": row["比赛ID"],
                "match_no": str(row["场次编号"]),
                "stage_order": int(row["阶段顺序"]),
                "round": row["轮次"],
                "stage": "knockout",
                "home_team": row["名义主队中文"],
                "home_team_en": row["名义主队英文"],
                "home_code": row["主队代码"],
                "away_team": row["名义客队中文"],
                "away_team_en": row["名义客队英文"],
                "away_code": row["客队代码"],
                "kickoff_local": row["当地开球时间"],
                "local_date": row["当地日期"],
                "local_time": row["当地时间"],
                "local_timezone": row["当地时区"],
                "kickoff_bjt": row["北京时间"],
                "beijing_date": row["北京时间"][:10],
                "beijing_time": row["北京时间"][11:16],
                "beijing_timezone": "Asia/Shanghai",
                "status": row["比赛状态"],
                "winner_advances_to": row["胜者晋级比赛ID"],
                "note": row.get("备注", ""),
            }
        )
    return out


def latest_standings() -> list[dict[str, Any]]:
    candidates = sorted(ROOT.glob("20*/standings_*.json"), reverse=True)
    for path in candidates:
        try:
            return read_json(path)
        except Exception:
            continue
    return []


def team_key(name: str) -> str:
    name = TEAM_ALIAS.get(name, name)
    return re.sub(r"\s+", "", name).lower()


def standings_lookup() -> dict[str, dict[str, Any]]:
    lookup: dict[str, dict[str, Any]] = {}
    for group in latest_standings():
        for row in group.get("rows", []):
            copied = dict(row)
            copied["group"] = group.get("group", copied.get("group", ""))
            lookup[team_key(str(row.get("team", "")))] = copied
    return lookup


def qualified_team_stats(schedule: list[dict[str, Any]]) -> list[dict[str, Any]]:
    first_round = [item for item in schedule if item["stage_order"] == 1 and not item["home_team"].startswith("待定")]
    team_order: list[tuple[str, str, str]] = []
    for item in first_round:
        team_order.append((item["home_team"], item["home_code"], item["game_id"]))
        team_order.append((item["away_team"], item["away_code"], item["game_id"]))

    lookup = standings_lookup()
    stats = []
    seen = set()
    for name, code, game_id in team_order:
        if name in seen:
            continue
        seen.add(name)
        row = lookup.get(team_key(name), {})
        rank = int(float(row.get("rank", 0) or 0))
        if rank == 1:
            path = "小组第一"
        elif rank == 2:
            path = "小组第二"
        elif rank == 3:
            path = "最佳第三名"
        else:
            path = "淘汰赛席位"
        stats.append(
            {
                "team_name": name,
                "team_code": code,
                "group": row.get("group", ""),
                "qualifying_path": path,
                "played": int(float(row.get("played", 0) or 0)),
                "wins": int(float(row.get("wins", 0) or 0)),
                "draws": int(float(row.get("draws", 0) or 0)),
                "losses": int(float(row.get("losses", 0) or 0)),
                "goals_for": int(float(row.get("gf", 0) or 0)),
                "goals_against": int(float(row.get("ga", 0) or 0)),
                "goal_difference": int(float(row.get("gd", 0) or 0)),
                "points": int(float(row.get("points", 0) or 0)),
                "first_knockout_match_id": game_id,
            }
        )
    return stats


def dt_bjt(value: str) -> datetime:
    cleaned = value.replace(" +08:00", "")
    return datetime.strptime(cleaned, "%Y-%m-%d %H:%M:%S")


def future_matches(schedule: list[dict[str, Any]], start: datetime, days: int = 3) -> list[dict[str, Any]]:
    end = start + timedelta(days=days)
    matches = [
        item
        for item in schedule
        if start <= dt_bjt(item["kickoff_bjt"]) < end and not item["home_team"].startswith("待定")
    ]
    return sorted(matches, key=lambda item: (item["beijing_date"], item["beijing_time"], int(item["match_no"])))


def current_bjt() -> datetime:
    override = os.environ.get("WORLD_CUP_NOW_BJT", "").strip()
    if override:
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y%m%d%H%M", "%Y%m%d"):
            try:
                return datetime.strptime(override, fmt)
            except ValueError:
                continue
        raise ValueError(f"Invalid WORLD_CUP_NOW_BJT: {override}")
    return datetime.now(timezone(timedelta(hours=8))).replace(tzinfo=None)


def settled_match_keys() -> set[str]:
    keys: set[str] = set()
    for path in DATA_DIR.glob("knockout_results_*.json"):
        try:
            payload = read_json(path)
        except Exception:
            continue
        for item in payload.get("matches", []):
            game_id = str(item.get("game_id", "")).strip()
            match_no = str(item.get("match_no", "")).strip().lstrip("0")
            if game_id:
                keys.add(f"game:{game_id}")
            if match_no:
                keys.add(f"no:{match_no}")
    return keys


def is_settled_match(item: dict[str, Any], settled: set[str]) -> bool:
    game_id = str(item.get("game_id", "")).strip()
    match_no = str(item.get("match_no", "")).strip().lstrip("0")
    return f"game:{game_id}" in settled or f"no:{match_no}" in settled


def rolling_future_matches(schedule: list[dict[str, Any]], days: int = 3) -> list[dict[str, Any]]:
    start = current_bjt()
    end = start + timedelta(days=days)
    settled = settled_match_keys()
    matches = [
        item
        for item in schedule
        if start <= dt_bjt(item["kickoff_bjt"]) < end
        and not item["home_team"].startswith("待定")
        and not is_settled_match(item, settled)
    ]
    return sorted(matches, key=lambda item: (item["beijing_date"], item["beijing_time"], int(item["match_no"])))


def round_loss_path(round_name: str, team: str) -> str:
    if round_name == "半决赛":
        return f"{team} 输球进入三四名决赛"
    if round_name == "决赛":
        return f"{team} 输球获得亚军"
    if round_name == "季军赛":
        return f"{team} 输球获得第四名"
    return f"{team} 输球即出局"


def next_opponent_source(schedule: list[dict[str, Any]], next_id: str, current_game_id: str) -> str:
    next_match = next((item for item in schedule if item["game_id"] == next_id), None)
    if not next_match:
        return "后续对手待定"
    sources = []
    for side in ("home_team", "away_team"):
        text = str(next_match[side])
        matched = re.search(r"(\d{8})", text)
        if matched and matched.group(1) != current_game_id:
            source_match = next((item for item in schedule if item["game_id"] == matched.group(1)), None)
            if source_match:
                sources.append(f"{source_match['home_team']} / {source_match['away_team']} 胜者")
    return "；".join(sources) if sources else "下一轮对手待定"


def resolved_next_opponent_source(schedule: list[dict[str, Any]], next_id: str, current_game_id: str) -> str:
    next_match = next((item for item in schedule if item["game_id"] == next_id), None)
    if not next_match:
        return "后续对手待定"
    winners = knockout_winner_lookup()
    sources = []
    for side in ("home_team", "away_team"):
        text = str(next_match[side])
        matched = re.search(r"(\d{8})", text)
        if matched and matched.group(1) != current_game_id:
            game_id = matched.group(1)
            source_match = next((item for item in schedule if item["game_id"] == game_id), None)
            if source_match:
                if game_id in winners:
                    sources.append(winners[game_id])
                else:
                    home = resolved_team_name(str(source_match["home_team"]), winners)
                    away = resolved_team_name(str(source_match["away_team"]), winners)
                    sources.append(f"{home} / {away} 胜者")
    return " / ".join(sources) if sources else "下一轮对手待定"


def prediction_model_v3() -> dict[str, Any]:
    return {
        "version": "v3-knockout-elo-ev-halftime",
        "model_sources": [
            "2026世界杯淘汰赛赛程表_中文版_Codex.md",
            "2026世界杯淘汰赛赛程表_中文版_Codex.xlsx",
            "world_cup_knockout_model_optimization_codex.md",
            "SHAREABLE_PREDICTION_REVIEW_LOGIC.md",
            "1.md",
        ],
        "mandatory_checks": [
            "赛程 Markdown 为主源，Excel 为一致性校验",
            "主客队、北京时间、当地时间、晋级路径必须逐项校验",
            "主模型和玄学面板彻底分离，玄学不进入综合评分",
            "输出必须含 90分钟胜平负、半场方向、最终晋级、比分 EV",
        ],
        "weights": {
            "elo_strength": 0.22,
            "group_stage_form": 0.18,
            "attack_defense_xg": 0.18,
            "odds_implied_market": 0.16,
            "knockout_script": 0.12,
            "injury_discipline": 0.06,
            "schedule_rest_travel": 0.04,
            "review_bias_correction": 0.04,
        },
        "ev_rules": [
            "EV = model_probability * market_decimal_odds - 1",
            "无真实赔率时只输出模型 EV 倾向，不给出下注指令",
            "比分 EV 以 Top6 覆盖为主，冷门比分单独标注，不覆盖主线",
        ],
        "data_pipeline": [
            "体彩实时赔率页背后的 getMatchCalculatorV1 接口为第一优先级赔率源，写入 had/ttg/hhad/crs/hafu。",
            "had/hhad 负责胜平负和让球强弱，ttg 负责总进球区间，crs 负责比分池，hafu 负责半场/全场节奏。",
            "worldcup26.ir/get/games 用作赛程和赛果兜底校验，不能覆盖已人工确认的赛程主客队。",
            "football-data.org、api-football、TheSportsDB、SerpApi 只作为球队资料、伤停、历史和新闻补充；缺数据时降权，不编造首发。",
        ],
        "parlay_policy": PARLAY_POLICY,
    }


def first_prediction(schedule: list[dict[str, Any]]) -> dict[str, Any]:
    match = next(item for item in schedule if item["game_id"] == FIRST_GAME_ID)
    home = match["home_team"]
    away = match["away_team"]
    next_source = next_opponent_source(schedule, match["winner_advances_to"], match["game_id"])
    return {
        "game_id": match["game_id"],
        "match_no": match["match_no"],
        "round": match["round"],
        "stage": "knockout",
        "home_team": home,
        "home_code": match["home_code"],
        "away_team": away,
        "away_code": match["away_code"],
        "kickoff_bjt": match["kickoff_bjt"],
        "kickoff_local": match["kickoff_local"],
        "beijing_date": match["beijing_date"],
        "model_version": "v3-knockout-elo-ev-halftime",
        "elo": {
            "home": 1622,
            "away": 1708,
            "diff_home_minus_away": -86,
            "elo_note": "加拿大综合强度略高，南非依靠低节奏和定位球保留冷门路径。",
        },
        "expected_goals": {"home_xg": 1.02, "away_xg": 1.16, "total_xg": 2.18},
        "half_time": {
            "direction": "平",
            "main_score": "0-0",
            "backup_score": "0-1",
            "probabilities": {"home": 0.22, "draw": 0.51, "away": 0.27},
        },
        "full_time_90": {
            "direction": "平",
            "probabilities": {"home": 0.28, "draw": 0.34, "away": 0.38},
            "double_chance": "加拿大不败 X2",
        },
        "direction_90min": "draw",
        "direction_text": "90分钟平局保护",
        "main_score": "1-1",
        "safe_scores": ["0-1", "1-2"],
        "draw_protection_scores": ["0-0", "1-1"],
        "upset_score": "1-0",
        "goals_range": "1-2球",
        "home_advance_probability": 0.46,
        "away_advance_probability": 0.54,
        "advance_pick": away,
        "advance_market": {"home": 0.46, "away": 0.54, "pick": away},
        "extra_time_penalty_script": "90分钟1-1，加拿大加时/点球小优。",
        "score_confidence": 0.61,
        "direction_confidence": 0.56,
        "data_completeness": 0.86,
        "team_profiles": {
            "home": [
                "小组赛三场：1胜1平1负、进2失3。能出线靠的是防守韧性，不是持续压制能力。",
                "已确认末轮 1-0 韩国：这说明南非在必须拿分时能把比赛压低，并通过单球解决问题。",
                "对捷克打成 1-1，对墨西哥这种小组最强档吃亏，说明南非遇到高压边路时守得住一段时间，但抗压上限有限。",
            ],
            "away": [
                "小组赛三场：1胜1平1负、进8失3。进攻爆发力明显强于南非，但防线不是零风险。",
                "已确认 6-0 卡塔尔和 1-2 瑞士：加拿大能打穿弱队，也会在强队对抗里丢球，不能简单按大胜思路处理。",
                "另一个平局样本把加拿大拉回现实：对抗更接近时，90分钟赢球并不稳，晋级优势更多体现在后程和点球前的深度。",
            ]
        },
        "game_plan": {
            "home": "南非：先守中路和二点球，争取半场不落后，靠定位球或反击保留 1-0 / 1-1 路径。",
            "away": "加拿大：前60分钟避免被拖进纯身体战，后30分钟用边路和换人把 0-1 / 1-2 路径打开。",
        },
        "model_insights": [
            {"title": "统一结论", "content": "这场不是“加拿大90分钟稳胜”。统一口径是：半场偏 0-0，90分钟防 1-1，最终晋级略偏加拿大。"},
            {"title": "为什么不追加拿大大胜", "content": "加拿大进攻更强，但南非已经证明能把必须拿分的比赛压成单球局；淘汰赛首战也会天然降低冒险程度。"},
            {"title": "为什么仍选加拿大晋级", "content": "加拿大有更好的进攻上限和后程换人空间。只要不被南非定位球先手，比赛越往后，加拿大越容易拿到决定性机会。"},
            {"title": "比分怎么用", "content": "主比分只看 1-1。保守备选是 0-1，开放备选是 1-2。南非冷门只保留 1-0，不再给多条互相冲突路线。"},
            {"title": "临场修正", "content": "若加拿大早段边路推进有效，0-1 升权；若南非定位球先手，比赛转向 1-0 / 1-1。"}
        ],
        "match_plan": [
            {"phase": "0-30分钟", "home": "南非优先压低节奏，减少中场身后空间。", "away": "加拿大试探边路，不急于把阵型压死。"},
            {"phase": "30-60分钟", "home": "若仍是平局，南非会接受慢节奏并等待定位球。", "away": "加拿大需要提高传中和肋部冲击频率。"},
            {"phase": "60-90分钟", "home": "南非体能下降后更依赖反击和死球。", "away": "加拿大晋级倾向在后程增强，0-1/1-2脚本开始升权。"},
            {"phase": "加时/点球", "home": "若拖到点球，南非冷门概率上升。", "away": "加拿大整体深度略优，但必须避免点球变成纯随机。"}
        ],
        "triggers": [
            "加拿大先入球：主比分从 1-1 下修为 0-1，开放尾部升到 1-2。",
            "南非先入球：比赛进入 1-0 冷门保护，同时加拿大反扑会带来 1-1。",
            "半场 0-0：继续保留 1-1 主线，加拿大晋级不等同于 90分钟客胜。",
            "早段黄牌偏多：总进球下修，半场平局和小比分权重提高。"
        ],
        "ev_table": [
            {"market": "90分钟平局", "model_probability": 0.34, "fair_odds": 2.94, "ev_signal": "+0.04", "note": "保护主线"},
            {"market": "加拿大晋级", "model_probability": 0.54, "fair_odds": 1.85, "ev_signal": "+0.06", "note": "最终结果优于90分钟客胜"},
            {"market": "半场平局", "model_probability": 0.51, "fair_odds": 1.96, "ev_signal": "+0.03", "note": "淘汰赛首战慢热"},
            {"market": "总进球1-2", "model_probability": 0.47, "fair_odds": 2.13, "ev_signal": "+0.02", "note": "低比分主仓"},
        ],
        "score_matrix_top": [
            {"score": "1-1", "probability": 0.176, "ev_signal": "+0.03", "label": "主比分"},
            {"score": "0-1", "probability": 0.142, "ev_signal": "+0.04", "label": "加拿大小胜"},
            {"score": "1-2", "probability": 0.116, "ev_signal": "+0.01", "label": "尾部上修"},
            {"score": "0-0", "probability": 0.104, "ev_signal": "+0.02", "label": "平局保护"},
            {"score": "1-0", "probability": 0.092, "ev_signal": "-0.01", "label": "南非冷门"},
            {"score": "2-1", "probability": 0.074, "ev_signal": "-0.02", "label": "南非开放冷门"},
        ],
        "scenario": {
            "winner_advances_to_match_id": match["winner_advances_to"],
            "next_opponent_source": next_source,
            "home_win_script": f"{home} 晋级后下轮对阵 {next_source}",
            "away_win_script": f"{away} 晋级后下轮对阵 {next_source}",
            "home_loss_script": round_loss_path(match["round"], home),
            "away_loss_script": round_loss_path(match["round"], away),
        },
        "analysis": [
            "淘汰赛首场优先拆分 90分钟赛果和最终晋级，不把加拿大晋级倾向等同于90分钟客胜。",
            "加拿大 Elo、进攻效率和小组赛净胜能力略好，但南非低节奏能力会压低比赛总进球。",
            "半场主线是 0-0；若加拿大边路早开局，半场 0-1 会提前触发 1-2 尾部脚本。",
        ],
        "risks": [
            "淘汰赛首战保守程度可能高于小组赛，比分仓位不宜重。",
            "若加拿大边路早早打穿，0-1 会升级为 1-2。",
            "若南非定位球先手，1-0 冷门路径成立。",
        ],
    }


ADDITIONAL_MATCH_MODELS: dict[str, dict[str, Any]] = {
    "53452557": {
        "model_version": "v3-knockout-elo-ev-halftime",
        "half_time": "巴西 1-0 日本",
        "direction_text": "巴西90分钟胜",
        "main_score": "2-1",
        "backup_scores": ["2-0", "1-1"],
        "upset_score": "1-2",
        "goals_range": "2-3球",
        "advance_pick": "巴西",
        "advance_probability_text": "巴西 62% / 日本 38%",
        "next_opponent_note": "胜者下轮对阵 科特迪瓦 / 挪威 胜者",
        "core_view": "巴西纸面和淘汰赛经验占优，日本有反击质量，但更像能制造进球而不是掀翻主线。",
        "team_reading": [
            "巴西小组赛进攻效率稳定，面对中低位防守时仍有单点破局能力。",
            "日本小组赛韧性强，适合把比赛拖进1球差，但防线承压时间过长会暴露身后空间。",
            "本场合理打法：巴西先抢节奏和禁区前沿，日本守住前30分钟后打反击。"
        ],
        "model_reasoning": [
            "主线不走大胜：日本具备把比分压住的能力。",
            "巴西胜面来自前场个人能力和淘汰赛控场经验。",
            "若半场仍是0-0，1-1保护升权；若巴西先入球，2-0/2-1成为主路径。"
        ],
        "mystic_summary": [
            "梅花/六爻：主队气更足，巴西小优。",
            "五行/飞星：黄绿木火有进攻象，但日本白蓝金水能守，比分不宜过大。",
            "玄学赛果：巴西2-1，防1-1。"
        ],
    },
    "53452541": {
        "model_version": "v3-knockout-elo-ev-halftime",
        "half_time": "德国 1-0 巴拉圭",
        "direction_text": "德国90分钟胜",
        "main_score": "2-0",
        "backup_scores": ["2-1", "1-0"],
        "upset_score": "1-1",
        "goals_range": "2-3球",
        "advance_pick": "德国",
        "advance_probability_text": "德国 68% / 巴拉圭 32%",
        "next_opponent_note": "胜者下轮对阵 法国 / 瑞典 胜者",
        "core_view": "德国小组赛火力和压迫质量更强，巴拉圭能守但反击持续性不足，德国常规时间解决概率最高。",
        "team_reading": [
            "德国小组赛进球能力强，进入淘汰赛后仍有主动控场空间。",
            "巴拉圭以第三名路径出线，防守优先，但如果先丢球很难长时间对攻。",
            "本场合理打法：德国前压制造早球，巴拉圭先守低位并等待定位球。"
        ],
        "model_reasoning": [
            "德国优势不是单纯纸面强，而是能持续制造射门和二次进攻。",
            "巴拉圭冷门条件只有一个：上半场守到0-0并通过定位球制造1-1。",
            "若德国30分钟前进球，2-0是主线；若久攻不下，1-0降速收场。"
        ],
        "mystic_summary": [
            "梅花/六爻：主队世爻旺，德国主动。",
            "奇门/五行：金水秩序感强，利德国控场。",
            "玄学赛果：德国2-0，防2-1。"
        ],
    },
    "53452547": {
        "model_version": "v3-knockout-elo-ev-halftime",
        "half_time": "荷兰 0-0 摩洛哥",
        "direction_text": "90分钟平局保护",
        "main_score": "1-1",
        "backup_scores": ["1-0", "0-1"],
        "upset_score": "0-2",
        "goals_range": "1-2球",
        "advance_pick": "荷兰",
        "advance_probability_text": "荷兰 53% / 摩洛哥 47%",
        "next_opponent_note": "胜者下轮对阵 南非 / 加拿大 胜者",
        "core_view": "这是四场里最接近的一场。荷兰整体结构更稳，摩洛哥反击和身体对抗有爆点，90分钟必须防平。",
        "team_reading": [
            "荷兰小组赛进攻数据好，但淘汰赛面对摩洛哥不会轻松打穿。",
            "摩洛哥韧性和转换速度足够制造冷门，尤其适合在下半场偷一个。",
            "本场合理打法：荷兰控球压节奏，摩洛哥守住中路后打身后。"
        ],
        "model_reasoning": [
            "荷兰晋级略优，但不等同于90分钟主胜。",
            "摩洛哥若先入球，比赛会转向0-1/0-2冷门线。",
            "主线是1-1；荷兰靠加时、点球或最后30分钟质量晋级。"
        ],
        "mystic_summary": [
            "梅花/六爻：主客相持，平局象重。",
            "奇门/飞星：伏吟偏慢，先守后变。",
            "玄学赛果：1-1，荷兰最终晋级。"
        ],
    },
    "53452561": {
        "model_version": "v3-knockout-elo-ev-halftime",
        "half_time": "科特迪瓦 0-0 挪威",
        "direction_text": "90分钟平局保护",
        "main_score": "1-1",
        "backup_scores": ["1-2", "0-1"],
        "upset_score": "2-1",
        "goals_range": "1-3球",
        "advance_pick": "挪威",
        "advance_probability_text": "科特迪瓦 45% / 挪威 55%",
        "next_opponent_note": "胜者下轮对阵 巴西 / 日本 胜者",
        "core_view": "科特迪瓦身体对抗和边路冲击能压低挪威节奏，但挪威禁区终结点更清楚，90分钟先防平，晋级略偏挪威。",
        "team_reading": [
            "科特迪瓦适合把比赛带入身体对抗和定位球节奏，能制造1球差冷门路径。",
            "挪威进攻结构更直接，禁区内终结质量高，但遇到高强度贴防时推进会断续。",
            "本场合理打法：科特迪瓦先压低中路空间，挪威避免急躁长传，等待边路传中和二点球。"
        ],
        "model_reasoning": [
            "6月30日三场验证了低比分与平局保护价值，尤其德国被拖成1-1后，强弱差不能直接放大成2球胜。",
            "挪威优势主要在终结点，不在全场压制；如果迟迟打不开，1-1会长期保持高权重。",
            "若挪威先入球，0-1/1-2升权；若科特迪瓦定位球先手，2-1冷门线成立。"
        ],
        "mystic_summary": [
            "梅花/六爻：客方后劲较足，先滞后通。",
            "五行/飞星：橙绿土木对抗红白金水，慢局象重。",
            "玄学赛果：1-1，挪威最终晋级。"
        ],
    },
    "53452543": {
        "model_version": "v3-knockout-elo-ev-halftime",
        "half_time": "法国 1-0 瑞典",
        "direction_text": "法国90分钟胜",
        "main_score": "2-1",
        "backup_scores": ["1-0", "1-1"],
        "upset_score": "0-1",
        "goals_range": "1-3球",
        "advance_pick": "法国",
        "advance_probability_text": "法国 64% / 瑞典 36%",
        "next_opponent_note": "胜者下轮对阵 德国 / 巴拉圭 胜者",
        "core_view": "法国单点爆破和阵容深度仍是主线，但瑞典定位球和防守纪律会把比分压住，主比分不追大胜。",
        "team_reading": [
            "法国在淘汰赛里有更好的前场个人能力和换人空间，能在僵持局里制造质量机会。",
            "瑞典低位防守和定位球威胁稳定，足够保留1-1或0-1冷门保护。",
            "本场合理打法：法国先抢半场领先，瑞典先守住禁区正面并争取定位球。"
        ],
        "model_reasoning": [
            "巴西2-1命中说明强方小胜仍要保留；德国1-1提醒模型不能低估北欧/南美低位防守的拖慢能力。",
            "法国胜面更清楚，但比分池要集中在1球差，避免把纸面优势直接翻译成3-0。",
            "若半场0-0，1-1显著升权；若法国早进球，2-1和2-0尾部升权。"
        ],
        "mystic_summary": [
            "梅花/六爻：主队动爻更旺，法国主动。",
            "奇门/五行：蓝白金水成势，利控制与尾段加速。",
            "玄学赛果：法国2-1，防1-1。"
        ],
    },
    "53452563": {
        "model_version": "v3-knockout-elo-ev-halftime",
        "half_time": "墨西哥 0-0 厄瓜多尔",
        "direction_text": "90分钟平局保护",
        "main_score": "1-1",
        "backup_scores": ["0-1", "1-0"],
        "upset_score": "1-2",
        "goals_range": "1-2球",
        "advance_pick": "厄瓜多尔",
        "advance_probability_text": "墨西哥 48% / 厄瓜多尔 52%",
        "next_opponent_note": "胜者下轮对阵 英格兰 / 刚果民主共和国 胜者",
        "core_view": "双方强弱差很窄，墨西哥主场气质和控场能撑住90分钟，厄瓜多尔转换和身体优势让最终晋级略优。",
        "team_reading": [
            "墨西哥节奏控制和比赛经验好，适合把淘汰赛拖进低比分。",
            "厄瓜多尔对抗和纵向推进更有爆点，尤其下半场能通过转换制造决定性机会。",
            "本场合理打法：墨西哥控球降速，厄瓜多尔守住肋部后打快速前插。"
        ],
        "model_reasoning": [
            "荷兰-摩洛哥1-1命中说明接近场次要坚持平局主线，不强行给胜负。",
            "厄瓜多尔晋级略优来自下半场冲击力，不等同于90分钟稳胜。",
            "若墨西哥先进球，1-0/1-1升权；若厄瓜多尔先入球，0-1/1-2成为主路径。"
        ],
        "mystic_summary": [
            "梅花/六爻：主客比和，平象重。",
            "飞星/五行：绿白相持，后段客方动能略强。",
            "玄学赛果：1-1，厄瓜多尔最终晋级。"
        ],
    },
}


ADDITIONAL_MATCH_DETAILS: dict[str, dict[str, Any]] = {
    "53452557": {
        "review_adjustment": "南非-加拿大复盘：低比分方向正确，但90分钟平局保护过重。加拿大补时绝杀说明淘汰赛强方即使打不开局，也要给90+阶段的边路、二点球和替补冲击保留小胜权重。本场巴西仍走胜，但比分从大胜压回2-1/2-0。",
        "score_matrix_top": [
            {"score": "2-1", "probability": 0.168, "ev_signal": "+0.04", "label": "主比分"},
            {"score": "2-0", "probability": 0.142, "ev_signal": "+0.03", "label": "巴西控场小胜"},
            {"score": "1-1", "probability": 0.118, "ev_signal": "+0.02", "label": "日本拖住节奏"},
            {"score": "1-0", "probability": 0.096, "ev_signal": "+0.01", "label": "低比分保护"},
            {"score": "1-2", "probability": 0.072, "ev_signal": "-0.01", "label": "日本冷门"},
            {"score": "3-1", "probability": 0.066, "ev_signal": "-0.02", "label": "尾部上修"},
        ],
        "ev_table": [
            {"market": "巴西90分钟胜", "model_probability": 0.54, "fair_odds": 1.85, "ev_signal": "+0.05", "note": "胜面保留，但不追穿盘大胜"},
            {"market": "总进球2-3", "model_probability": 0.51, "fair_odds": 1.96, "ev_signal": "+0.04", "note": "主比分池集中在2球和3球"},
            {"market": "半场巴西不败", "model_probability": 0.72, "fair_odds": 1.39, "ev_signal": "+0.02", "note": "日本前30分钟抗压是关键"},
            {"market": "日本进球", "model_probability": 0.46, "fair_odds": 2.17, "ev_signal": "+0.01", "note": "反击质量足够制造一个球"},
        ],
        "model_insights": [
            {"title": "统一结论", "content": "巴西90分钟胜是主线，但不是碾压局；日本的转换能力要求比分池保留1-1和2-1。"},
            {"title": "复盘校正", "content": "昨日加拿大补时进球提醒模型：强方久攻不下时不能简单降为平局，要保留90分钟末段小胜。"},
            {"title": "比分怎么用", "content": "主比分2-1，保守备选2-0，若半场0-0则1-1升权；巴西早进球才上修3-1。"},
        ],
        "triggers": [
            "巴西30分钟前进球：2-0、3-1升权，日本被迫前压后空间会变大。",
            "半场0-0：1-1保护升权，巴西胜从常规胜转为尾段小胜。",
            "日本先进球：比赛转向1-1/1-2冷门线，巴西需要尽快提高射门频率。",
        ],
    },
    "53452541": {
        "review_adjustment": "南非-加拿大复盘：淘汰赛首战的低比分、防守耐心和补时决胜都得到验证。德国优势更明显，但巴拉圭的低位防守会把比赛压成1-0/2-0区间，冷门主要来自定位球1-1。",
        "score_matrix_top": [
            {"score": "2-0", "probability": 0.184, "ev_signal": "+0.05", "label": "主比分"},
            {"score": "1-0", "probability": 0.146, "ev_signal": "+0.03", "label": "久攻小胜"},
            {"score": "2-1", "probability": 0.122, "ev_signal": "+0.02", "label": "定位球丢球"},
            {"score": "1-1", "probability": 0.092, "ev_signal": "+0.01", "label": "冷门保护"},
            {"score": "3-0", "probability": 0.081, "ev_signal": "-0.01", "label": "早球放大"},
            {"score": "3-1", "probability": 0.064, "ev_signal": "-0.02", "label": "尾部上修"},
        ],
        "ev_table": [
            {"market": "德国90分钟胜", "model_probability": 0.61, "fair_odds": 1.64, "ev_signal": "+0.06", "note": "今日三场里最明确的方向"},
            {"market": "德国零封", "model_probability": 0.45, "fair_odds": 2.22, "ev_signal": "+0.03", "note": "巴拉圭运动战持续性不足"},
            {"market": "总进球2-3", "model_probability": 0.49, "fair_odds": 2.04, "ev_signal": "+0.02", "note": "不追4球以上"},
            {"market": "半场德国领先", "model_probability": 0.43, "fair_odds": 2.33, "ev_signal": "+0.01", "note": "早段压迫质量决定上限"},
        ],
        "model_insights": [
            {"title": "统一结论", "content": "德国胜面清楚，主线2-0；巴拉圭能守，但如果先丢球，反扑质量不足。"},
            {"title": "复盘校正", "content": "昨日低比分和补时绝杀说明淘汰赛不能只按实力差拉大比分，德国胜也要优先用1-0/2-0表达。"},
            {"title": "比分怎么用", "content": "2-0为主，1-0是久攻不下的保守脚本，2-1来自巴拉圭定位球或二点球。"},
        ],
        "triggers": [
            "德国30分钟前进球：2-0主线加强，3-0进入尾部。",
            "半场0-0：1-0升权，巴拉圭1-1冷门保护同步上升。",
            "巴拉圭定位球威胁连续出现：2-1和1-1权重提高。",
        ],
    },
    "53452547": {
        "review_adjustment": "南非-加拿大复盘：平局保护有价值，但不能忽视补时阶段强侧或阵容深度更好一方的小胜。荷兰-摩洛哥强弱差更窄，因此仍以1-1为主，但把0-1/1-0同时放进核心比分池。",
        "score_matrix_top": [
            {"score": "1-1", "probability": 0.176, "ev_signal": "+0.05", "label": "主比分"},
            {"score": "1-0", "probability": 0.126, "ev_signal": "+0.03", "label": "荷兰尾段小胜"},
            {"score": "0-1", "probability": 0.118, "ev_signal": "+0.03", "label": "摩洛哥反击"},
            {"score": "0-0", "probability": 0.101, "ev_signal": "+0.02", "label": "低节奏锁局"},
            {"score": "2-1", "probability": 0.084, "ev_signal": "-0.01", "label": "荷兰开放胜"},
            {"score": "0-2", "probability": 0.061, "ev_signal": "-0.01", "label": "摩洛哥冷门放大"},
        ],
        "ev_table": [
            {"market": "90分钟平局", "model_probability": 0.35, "fair_odds": 2.86, "ev_signal": "+0.05", "note": "双方都能接受先不犯错"},
            {"market": "总进球1-2", "model_probability": 0.52, "fair_odds": 1.92, "ev_signal": "+0.04", "note": "低比分主线最稳"},
            {"market": "荷兰晋级", "model_probability": 0.53, "fair_odds": 1.89, "ev_signal": "+0.02", "note": "晋级略优，不等于90分钟主胜"},
            {"market": "双方进球", "model_probability": 0.47, "fair_odds": 2.13, "ev_signal": "+0.01", "note": "1-1是最高权重比分"},
        ],
        "model_insights": [
            {"title": "统一结论", "content": "这场不适合强行分胜负。90分钟1-1优先，最终晋级荷兰略占优。"},
            {"title": "复盘校正", "content": "昨日南非-加拿大证明低比分和平局保护有效，但补时小胜必须保留，所以1-0和0-1都进入核心池。"},
            {"title": "比分怎么用", "content": "主比分1-1；荷兰尾段质量对应1-0，摩洛哥转换效率对应0-1，0-0只作为节奏极慢保护。"},
        ],
        "triggers": [
            "半场0-0：1-1和0-0继续升权，荷兰晋级倾向不等于90分钟赢球。",
            "摩洛哥先进球：0-1/0-2冷门线升权，荷兰会被迫打开边路。",
            "荷兰60分钟后换人提升压迫：1-0、2-1升权。",
        ],
    },
    "53452561": {
        "review_adjustment": "6月30日复盘：巴西2-1与荷兰1-1命中，说明强方小胜和平局保护都有效；德国1-1偏离主线，暴露低位防守和定位球冷门权重仍不足。本场挪威只给晋级小优，90分钟先防1-1。",
        "score_matrix_top": [
            {"score": "1-1", "probability": 0.164, "ev_signal": "+0.04", "label": "主比分"},
            {"score": "0-1", "probability": 0.132, "ev_signal": "+0.03", "label": "挪威小胜"},
            {"score": "1-2", "probability": 0.108, "ev_signal": "+0.02", "label": "后段上修"},
            {"score": "0-0", "probability": 0.096, "ev_signal": "+0.02", "label": "慢热锁局"},
            {"score": "2-1", "probability": 0.082, "ev_signal": "-0.01", "label": "科特迪瓦定位球冷门"},
            {"score": "2-2", "probability": 0.061, "ev_signal": "-0.02", "label": "开放尾部"},
        ],
        "ev_table": [
            {"market": "90分钟平局", "model_probability": 0.34, "fair_odds": 2.94, "ev_signal": "+0.04", "note": "双方都不适合早段冒险"},
            {"market": "挪威晋级", "model_probability": 0.55, "fair_odds": 1.82, "ev_signal": "+0.03", "note": "终结点优势更适合加时/点球前兑现"},
            {"market": "总进球1-3", "model_probability": 0.61, "fair_odds": 1.64, "ev_signal": "+0.03", "note": "不追大球"},
            {"market": "双方进球", "model_probability": 0.48, "fair_odds": 2.08, "ev_signal": "+0.01", "note": "身体对抗和定位球都能制造进球"},
        ],
        "model_insights": [
            {"title": "统一结论", "content": "90分钟1-1优先，最终晋级略偏挪威；不把挪威终结优势直接翻译成常规时间稳胜。"},
            {"title": "复盘校正", "content": "德国1-1提醒模型：遇到低位和定位球强队，强方胜面要降一档，用平局保护承接。"},
            {"title": "比分怎么用", "content": "主比分1-1，挪威小胜看0-1/1-2，科特迪瓦冷门只保留2-1。"},
        ],
        "triggers": [
            "半场0-0：1-1和0-0继续升权，挪威晋级不等于90分钟胜。",
            "挪威先进球：0-1、1-2升权，科特迪瓦会提高边路冲击。",
            "科特迪瓦定位球连续制造威胁：2-1冷门和1-1主线同步升权。",
        ],
    },
    "53452543": {
        "review_adjustment": "6月30日复盘：巴西2-1说明强方一球小胜仍是淘汰赛主线；德国1-1说明不能低估纪律型防守的拖慢能力。本场法国仍走胜，但比分收窄到2-1/1-0，并保留1-1保护。",
        "score_matrix_top": [
            {"score": "2-1", "probability": 0.172, "ev_signal": "+0.05", "label": "主比分"},
            {"score": "1-0", "probability": 0.138, "ev_signal": "+0.03", "label": "法国小胜"},
            {"score": "1-1", "probability": 0.116, "ev_signal": "+0.02", "label": "瑞典拖住"},
            {"score": "2-0", "probability": 0.102, "ev_signal": "+0.01", "label": "法国控场"},
            {"score": "0-1", "probability": 0.066, "ev_signal": "-0.01", "label": "瑞典定位球冷门"},
            {"score": "3-1", "probability": 0.058, "ev_signal": "-0.02", "label": "早球放大"},
        ],
        "ev_table": [
            {"market": "法国90分钟胜", "model_probability": 0.57, "fair_odds": 1.75, "ev_signal": "+0.05", "note": "胜面明确但只按小胜处理"},
            {"market": "总进球2-3", "model_probability": 0.52, "fair_odds": 1.92, "ev_signal": "+0.04", "note": "主比分集中在2-1/1-1"},
            {"market": "瑞典进球", "model_probability": 0.44, "fair_odds": 2.27, "ev_signal": "+0.02", "note": "定位球和二点球不能忽视"},
            {"market": "半场法国不败", "model_probability": 0.73, "fair_odds": 1.37, "ev_signal": "+0.01", "note": "法国开局控制力更好"},
        ],
        "model_insights": [
            {"title": "统一结论", "content": "法国90分钟胜是主线，主比分2-1；瑞典的防守纪律要求保留1-0和1-1。"},
            {"title": "复盘校正", "content": "德国被巴拉圭拖平后，模型对纪律型防守的平局保护上调。"},
            {"title": "比分怎么用", "content": "2-1为主，1-0是低节奏小胜，1-1是半场打不开时的保护。"},
        ],
        "triggers": [
            "法国30分钟前进球：2-0、3-1升权，瑞典必须提前打开阵型。",
            "半场0-0：1-1和1-0升权，法国胜面从压制胜转为尾段小胜。",
            "瑞典定位球质量持续出现：2-1和0-1冷门保护同步上升。",
        ],
    },
    "53452563": {
        "review_adjustment": "6月30日复盘：荷兰-摩洛哥1-1完全符合接近场次的平局保护逻辑。本场墨西哥-厄瓜多尔同属窄差对抗，主线继续用1-1表达，晋级小优交给厄瓜多尔的后段冲击。",
        "score_matrix_top": [
            {"score": "1-1", "probability": 0.181, "ev_signal": "+0.05", "label": "主比分"},
            {"score": "0-1", "probability": 0.124, "ev_signal": "+0.03", "label": "厄瓜多尔小胜"},
            {"score": "1-0", "probability": 0.119, "ev_signal": "+0.03", "label": "墨西哥控场小胜"},
            {"score": "0-0", "probability": 0.104, "ev_signal": "+0.02", "label": "慢节奏锁局"},
            {"score": "1-2", "probability": 0.079, "ev_signal": "-0.01", "label": "厄瓜多尔后段放大"},
            {"score": "2-1", "probability": 0.072, "ev_signal": "-0.01", "label": "墨西哥开放胜"},
        ],
        "ev_table": [
            {"market": "90分钟平局", "model_probability": 0.36, "fair_odds": 2.78, "ev_signal": "+0.05", "note": "窄差对抗优先防平"},
            {"market": "总进球1-2", "model_probability": 0.54, "fair_odds": 1.85, "ev_signal": "+0.04", "note": "低比分最稳"},
            {"market": "厄瓜多尔晋级", "model_probability": 0.52, "fair_odds": 1.92, "ev_signal": "+0.02", "note": "后段身体优势略好"},
            {"market": "双方进球", "model_probability": 0.49, "fair_odds": 2.04, "ev_signal": "+0.01", "note": "1-1是最高权重比分"},
        ],
        "model_insights": [
            {"title": "统一结论", "content": "90分钟1-1优先，最终晋级略偏厄瓜多尔；两队都不具备稳定碾压对方的条件。"},
            {"title": "复盘校正", "content": "荷兰1-1命中后，接近场次继续把平局保护放在主线，而不是强行选胜负。"},
            {"title": "比分怎么用", "content": "1-1为主，0-1和1-0是两侧小胜保护，1-2只在厄瓜多尔先进球后上修。"},
        ],
        "triggers": [
            "半场0-0：1-1和0-0升权，比赛更可能进入加时/点球逻辑。",
            "厄瓜多尔先进球：0-1、1-2升权，墨西哥会被迫提高边路投入。",
            "墨西哥先进球：1-0、1-1升权，厄瓜多尔后段反扑会制造双方进球。"
        ],
    },
}


def additional_prediction(schedule: list[dict[str, Any]], game_id: str) -> dict[str, Any]:
    match = next(item for item in schedule if item["game_id"] == game_id)
    model = ADDITIONAL_MATCH_MODELS[game_id]
    details = ADDITIONAL_MATCH_DETAILS[game_id]
    return {
        **model,
        **details,
        "safe_scores": model["backup_scores"],
        "draw_protection_scores": ["0-0", "1-1"],
        "game_id": match["game_id"],
        "match_no": match["match_no"],
        "round": match["round"],
        "home_team": match["home_team"],
        "away_team": match["away_team"],
        "kickoff_bjt": match["kickoff_bjt"],
        "kickoff_local": match["kickoff_local"],
        "beijing_date": match["beijing_date"],
    }


def all_predictions(schedule: list[dict[str, Any]]) -> list[dict[str, Any]]:
    predictions = [first_prediction(schedule)]
    predictions.extend(additional_prediction(schedule, game_id) for game_id in PREDICTION_GAME_IDS[1:])
    return predictions


def mystic_panel() -> dict[str, Any]:
    return {
        "title": "玄学综参",
        "status": "仅作娱乐表达，不进入 Elo、EV、半场、胜平负和比分概率。",
        "local_time_rule": "起卦使用比赛当地时间：2026-06-28 15:00 -04:00。",
        "data_pack": [
            {"field": "主队", "value": "南非；现 SAFA 1991 年重组/成立；队名首字“南”按 9 画取数；球衣意象黄绿，对应土木。"},
            {"field": "客队", "value": "加拿大；加拿大足协源流 1912 年；队名首字“加”按 5 画取数；球衣意象红白，对应火金。"},
            {"field": "时间", "value": "当地 2026-06-28 15:00，按申时取 9；本场为淘汰赛首战，慢热和低比分权重提高。"},
            {"field": "方位", "value": "按中立场处理，名义主队为地盘/世爻，名义客队为天盘/应爻。"},
        ],
        "framework": [
            {"name": "梅花易数", "content": "结果：体卦乾金克用卦巽木，南非90分钟有抗衡。赛果预测：1-1，防南非1-0。"},
            {"name": "六爻世应", "content": "结果：初至三爻偏静，四至六爻客方动能增强。赛果预测：半场0-0，终局1-1，加拿大加时/点球更顺。"},
            {"name": "紫微流时", "content": "结果：申时金气利纪律和防守，客方迁移宫动。赛果预测：常规时间小比分，加拿大后段占优。"},
            {"name": "奇门遁甲", "content": "结果：伏吟象重，主客比和，节奏慢。赛果预测：90分钟平局优先，点球/加时分胜负。"},
            {"name": "干支五行", "content": "结果：火金并见，加拿大红白火金更得时，南非黄绿能守但破局不足。赛果预测：0-1 或 1-1。"},
            {"name": "九宫飞星", "content": "结果：小球象强，大球星不旺。赛果预测：总进球1-2球，重点0-0、1-1、0-1。"},
            {"name": "诸葛神数 / 米卦", "content": "结果：相持后动，先难后易。赛果预测：上半场僵持，下半场加拿大更接近进球。"},
            {"name": "塔罗辅助", "content": "结果：隐士-节制-命运之轮。赛果预测：谨慎开局、均衡中段、尾段或点球转折。"},
            {"name": "⚡民俗趣味", "content": "结果：黄绿主守，红白主冲。赛果预测：南非不易崩盘，加拿大不宜大胜。"},
        ],
        "summary": {
            "旺方": "均势偏客队后程；南非有90分钟抗衡象，加拿大有最终晋级象。",
            "宜": "小球，重点 0-0 / 1-1 / 0-1；若早球出现再上修 1-2。",
            "爆冷征兆": "南非定位球先手、早段黄牌压低节奏、半场0-0后进入点球随机区。",
            "玄学日运": "火金并见，先守后动；不宜追热大胜，宜防平局与低比分。",
        },
        "reading": "玄学结论与主模型一致保留平局保护，但玄学更强调南非90分钟不败阻力；最终晋级仍不覆盖主模型的加拿大小优。",
    }


def write_json(path: Path, data: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def render_score_matrix(rows: list[dict[str, Any]]) -> str:
    rows = compact_score_rows(rows)
    return "".join(
        "<tr>"
        f"<td>{esc(item['score'])}</td><td>{float(item['probability']) * 100:.1f}%</td>"
        f"<td>{esc(item.get('ev_signal', ''))}</td><td>{esc(item['label'])}</td></tr>"
        for item in rows
    )


def compact_score_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    safe_rows = [row for row in rows if str(row.get("ev_signal", "")).startswith("+")]
    upset_rows = [row for row in rows if "冷" in str(row.get("ev_signal", "")) or "冷" in str(row.get("label", ""))]
    selected: list[dict[str, Any]] = []
    seen: set[str] = set()
    for row in safe_rows[:3]:
        score = str(row.get("score", ""))
        if score and score not in seen:
            selected.append(row)
            seen.add(score)
    for row in upset_rows[:1]:
        score = str(row.get("score", ""))
        if score and score not in seen:
            selected.append(row)
            seen.add(score)
    if not selected:
        return rows[:4]
    return selected


def render_ev_rows(rows: list[dict[str, Any]]) -> str:
    return "".join(
        "<tr>"
        f"<td>{esc(item['market'])}</td><td>{float(item['model_probability']) * 100:.1f}%</td>"
        f"<td>{float(item['fair_odds']):.2f}</td><td>{esc(item['ev_signal'])}</td><td>{esc(item['note'])}</td></tr>"
        for item in rows
    )


def render_insight_cards(rows: list[dict[str, str]]) -> str:
    return "".join(
        f"<div class=\"panel\"><h3>{esc(item['title'])}</h3><p>{esc(item['content'])}</p></div>"
        for item in rows
    )


def render_match_plan(rows: list[dict[str, str]]) -> str:
    return "".join(
        "<tr>"
        f"<td>{esc(item['phase'])}</td><td>{esc(item['home'])}</td><td>{esc(item['away'])}</td></tr>"
        for item in rows
    )


def render_bullets(rows: list[str]) -> str:
    return "".join(f"<li>{esc(item)}</li>" for item in rows)


def render_review_lessons(review: dict[str, Any]) -> str:
    return "".join(f"<li>{esc(item)}</li>" for item in review.get("lessons", []))


def render_stats_rows(stats: list[dict[str, Any]]) -> str:
    rows = []
    for item in stats:
        rows.append(
            "<tr>"
            f"<td>{esc(item['team_name'])}</td><td>{esc(item['team_code'])}</td><td>{esc(item['group'])}</td>"
            f"<td>{esc(item['qualifying_path'])}</td><td>{item['wins']}</td><td>{item['draws']}</td><td>{item['losses']}</td>"
            f"<td>{item['goals_for']}</td><td>{item['goals_against']}</td><td>{item['goal_difference']}</td><td>{item['points']}</td></tr>"
        )
    return "".join(rows)


def render_upcoming_rows(matches: list[dict[str, Any]]) -> str:
    return "".join(
        "<tr>"
        f"<td>{esc(item['beijing_date'][5:])} {esc(item['beijing_time'])}</td>"
        f"<td>{esc(item['match_no'])}</td><td>{esc(item['round'])}</td>"
        f"<td>{esc(item['home_team'])} vs {esc(item['away_team'])}</td>"
        f"<td>{esc(next_opponent_source(GLOBAL_SCHEDULE, item['winner_advances_to'], item['game_id']))}</td></tr>"
        for item in matches
    )


def knockout_winner_lookup() -> dict[str, str]:
    winners: dict[str, str] = {}
    for path in sorted(DATA_DIR.glob("knockout_results_*.json")):
        try:
            payload = read_json(path)
        except Exception:
            continue
        for item in payload.get("matches", []):
            game_id = str(item.get("game_id", "")).strip()
            if not game_id:
                continue
            winner = str(item.get("advance_team") or item.get("winner_team") or item.get("penalty_winner") or "").strip()
            home_goals = item.get("home_goals")
            away_goals = item.get("away_goals")
            if not winner and home_goals is not None and away_goals is not None:
                if int(home_goals) > int(away_goals):
                    winner = str(item.get("home_team", "")).strip()
                elif int(home_goals) < int(away_goals):
                    winner = str(item.get("away_team", "")).strip()
            if winner:
                winners[game_id] = winner
    return winners


def resolved_team_name(team_name: str, winners: dict[str, str]) -> str:
    text = str(team_name)
    matched = re.search(r"(\d{8})", text)
    if matched:
        return winners.get(matched.group(1), text)
    return text


def future_round16_matches(schedule: list[dict[str, Any]]) -> list[dict[str, Any]]:
    winners = knockout_winner_lookup()
    settled = settled_match_keys()
    rows: list[dict[str, Any]] = []
    for item in schedule:
        try:
            match_no = int(str(item.get("match_no", "0")))
        except ValueError:
            continue
        if 89 <= match_no <= 96 and not is_settled_match(item, settled):
            copied = dict(item)
            copied["home_team_resolved"] = resolved_team_name(str(item.get("home_team", "")), winners)
            copied["away_team_resolved"] = resolved_team_name(str(item.get("away_team", "")), winners)
            rows.append(copied)
    return sorted(rows, key=lambda item: int(str(item["match_no"])))


def render_round16_schedule_rows(matches: list[dict[str, Any]]) -> str:
    if not matches:
        return '<tr><td colspan="4">暂无未来16强赛程。</td></tr>'
    return "".join(
        "<tr>"
        f"<td>{esc(item['beijing_date'][5:])} {esc(item['beijing_time'])}</td>"
        f"<td>{esc(item['match_no'])}</td>"
        f"<td>{esc(item['home_team_resolved'])} vs {esc(item['away_team_resolved'])}</td>"
        f"<td>{esc(resolved_next_opponent_source(GLOBAL_SCHEDULE, item['winner_advances_to'], item['game_id']))}</td>"
        "</tr>"
        for item in matches
    )


def round16_market_forecasts() -> list[dict[str, Any]]:
    path = DATA_DIR / "round16_hhad_predictions_20260704.json"
    if not path.exists():
        return []
    try:
        settled = settled_match_keys()
        return [
            item
            for item in read_json(path).get("matches", [])
            if f"game:{item.get('game_id', '')}" not in settled
            and f"no:{str(item.get('match_no', '')).lstrip('0')}" not in settled
        ]
    except Exception:
        return []


def render_round16_market_rows(rows: list[dict[str, Any]]) -> str:
    if not rows:
        return '<tr><td colspan="7">方向类市场已停止输出。</td></tr>'
    return "".join(
        "<tr>"
        f"<td>{esc(item.get('match_no', ''))}</td>"
        f"<td>{esc(item.get('kickoff_bjt', '').replace('2026-', '').replace(':00 +08:00', ''))}</td>"
        f"<td>{esc(item.get('home_team', ''))} vs {esc(item.get('away_team', ''))}</td>"
        f"<td>{esc(item.get('total_goals_pick', ''))}</td>"
        f"<td>{esc(item.get('total_goals_range', ''))}</td>"
        f"<td>{esc(item.get('total_goals_backup', ''))}</td>"
        f"<td>{esc(item.get('total_goals_reason', ''))}</td>"
        "</tr>"
        for item in rows
    )


def quarterfinal_matches(schedule: list[dict[str, Any]]) -> list[dict[str, Any]]:
    winners = knockout_winner_lookup()
    rows: list[dict[str, Any]] = []
    for item in schedule:
        try:
            match_no = int(str(item.get("match_no", "0")))
        except ValueError:
            continue
        if 97 <= match_no <= 100:
            copied = dict(item)
            copied["home_team_resolved"] = resolved_team_name(str(item.get("home_team", "")), winners)
            copied["away_team_resolved"] = resolved_team_name(str(item.get("away_team", "")), winners)
            rows.append(copied)
    return sorted(rows, key=lambda item: int(str(item["match_no"])))


def entry_status(team_name: str) -> str:
    return "已入围" if not str(team_name).startswith("待定") else "待产生"


def render_quarterfinal_entry_rows(matches: list[dict[str, Any]]) -> str:
    if not matches:
        return '<tr><td colspan="6">暂无8强入围信息。</td></tr>'
    rendered = []
    for item in matches:
        home = item.get("home_team_resolved", "")
        away = item.get("away_team_resolved", "")
        rendered.append(
            "<tr>"
            f"<td>{esc(item['beijing_date'][5:])} {esc(item['beijing_time'])}</td>"
            f"<td>{esc(item['match_no'])}</td>"
            f"<td>{esc(home)}</td>"
            f"<td>{esc(entry_status(home))}</td>"
            f"<td>{esc(away)}</td>"
            f"<td>{esc(entry_status(away))}</td>"
            "</tr>"
        )
    return "".join(rendered)


def total_goals_from_scores(item: dict[str, Any]) -> tuple[str, str, str]:
    scores: list[str] = []
    for key in ("main_score", "upset_score"):
        value = item.get(key)
        if value:
            scores.append(str(value))
    for key in ("backup_scores", "safe_scores", "draw_protection_scores"):
        values = item.get(key) or []
        if isinstance(values, list):
            scores.extend(str(value) for value in values if value)

    totals = [total for total in (score_total(score) for score in scores) if total is not None]
    if not totals:
        return "待定", "待定", "等待赔率/比分池刷新"

    counts: dict[int, int] = {}
    for total in totals:
        counts[total] = counts.get(total, 0) + 1
    ordered = sorted(counts.items(), key=lambda pair: (-pair[1], pair[0]))
    primary = ordered[0][0]
    backups = [str(total) for total, _ in ordered[1:3]]
    low = max(0, min(totals))
    high = max(totals)
    return f"{primary}球", f"{low}-{high}球", " / ".join(f"{value}球" for value in backups) if backups else "无"


def render_round16_total_goals_rows(rows: list[dict[str, Any]]) -> str:
    if not rows:
        return '<tr><td colspan="7">暂无16强总进球预测。</td></tr>'
    rendered = []
    for item in rows:
        inferred_pick, inferred_range, inferred_backup = total_goals_from_scores(item)
        pick = item.get("total_goals_pick") or inferred_pick
        goals_range = item.get("total_goals_range") or inferred_range
        backup = item.get("total_goals_backup") or inferred_backup
        reason = item.get("total_goals_reason") or item.get("reason", "")
        rendered.append(
            "<tr>"
            f"<td>{esc(item.get('match_no', ''))}</td>"
            f"<td>{esc(item.get('kickoff_bjt', '').replace('2026-', '').replace(':00 +08:00', ''))}</td>"
            f"<td>{esc(item.get('home_team', ''))} vs {esc(item.get('away_team', ''))}</td>"
            f"<td>{esc(pick)}</td>"
            f"<td>{esc(goals_range)}</td>"
            f"<td>{esc(backup)}</td>"
            f"<td>{esc(reason)}</td>"
            "</tr>"
        )
    return "".join(rendered)


def future_parlay_payload() -> dict[str, Any]:
    path = DATA_DIR / "round16_parlay_20260705.json"
    if not path.exists():
        return {}
    try:
        return read_json(path)
    except Exception:
        return {}


def render_future_parlay_group(group: dict[str, Any]) -> str:
    rows = "".join(
        "<tr>"
        f"<td>{esc(item.get('match_no', ''))}</td>"
        f"<td>{esc(item.get('match', ''))}</td>"
        f"<td>{esc(item.get('score_pick', ''))}</td>"
        f"<td>{esc(item.get('goals_pick', ''))}</td>"
        f"<td>{esc(item.get('half_full_pick', ''))}</td>"
        f"<td>{esc(item.get('ev_signal', ''))}</td>"
        f"<td>{esc(item.get('note', ''))}</td>"
        "</tr>"
        for item in group.get("legs", [])
    )
    return f"""
    <div class="card">
      <h3>{esc(group.get('title', '三串一'))}</h3>
      <p class="muted">{esc(group.get('summary', ''))}</p>
      <div class="tableWrap"><table><thead><tr><th>场次</th><th>比赛</th><th>比分稳胆</th><th>总进球</th><th>半全场</th><th>EV信号</th><th>说明</th></tr></thead><tbody>{rows}</tbody></table></div>
    </div>"""


def render_future_parlay_section(payload: dict[str, Any]) -> str:
    groups = payload.get("groups", [])
    if not groups:
        return ""
    body = "".join(render_future_parlay_group(group) for group in groups)
    return f"""
  <section class="section" id="futureParlay">
    <h2>未来串关</h2>
    <div class="card">
      <p><strong>{esc(payload.get('headline', '未赛场次三串一更新'))}</strong></p>
      <p>{esc(payload.get('summary', ''))}</p>
    </div>
    {body}
  </section>"""


def render_future_parlay_page(payload: dict[str, Any]) -> str:
    body = render_future_parlay_section(payload)
    if not body:
        body = '<section class="section"><h2>未来串关</h2><div class="card"><p>暂无可用串关。</p></div></section>'
    return f"""<!DOCTYPE html>
<html lang="zh-CN">
<head><meta charset="UTF-8"><meta name="viewport" content="width=device-width, initial-scale=1.0"><title>未来串关</title><style>{css()}</style></head>
<body>
<header>
  <h1>未来串关</h1>
  <nav><a href="../index.html">首页</a><a href="../20260706/">07-06预测</a><a href="../knockout/">淘汰赛日期</a></nav>
</header>
<main>
  {body}
</main>
</body>
</html>"""


def knockout_results() -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for path in sorted(DATA_DIR.glob("knockout_results_*.json")):
        try:
            payload = read_json(path)
        except Exception:
            continue
        date_label = str(payload.get("date") or path.stem.rsplit("_", 1)[-1])
        for item in payload.get("matches", []):
            home_goals = item.get("home_goals")
            away_goals = item.get("away_goals")
            if home_goals is None or away_goals is None:
                score_text = str(item.get("score", ""))
                matched = re.match(r"^\s*(\d+)\s*[-:：]\s*(\d+)\s*$", score_text)
                if matched:
                    home_goals = int(matched.group(1))
                    away_goals = int(matched.group(2))
            if home_goals is None or away_goals is None:
                result_text = "已赛"
            elif int(home_goals) > int(away_goals):
                result_text = f"{item.get('home_team', '')}胜"
            elif int(home_goals) < int(away_goals):
                result_text = f"{item.get('away_team', '')}胜"
            else:
                result_text = "90分钟平"
            rows.append(
                {
                    "date": date_label,
                    "date_text": f"{date_label[4:6]}-{date_label[6:8]}" if len(date_label) == 8 else date_label,
                    "match_no": str(item.get("match_no", "")),
                    "home_team": item.get("home_team", ""),
                    "away_team": item.get("away_team", ""),
                    "score": item.get("score", ""),
                    "result": result_text,
                    "prediction_main_score": item.get("prediction_main_score", ""),
                    "review": item.get("review", ""),
                    "half_time_score": item.get("half_time_score", ""),
                    "half_time": item.get("half_time", ""),
                    "half_score": item.get("half_score", ""),
                    "half_home_goals": item.get("half_home_goals"),
                    "half_away_goals": item.get("half_away_goals"),
                }
            )
    return sorted(rows, key=lambda item: (item["date"], int(str(item["match_no"] or "0"))))


def render_knockout_result_rows(rows: list[dict[str, Any]]) -> str:
    if not rows:
        return '<tr><td colspan="7">暂无淘汰赛赛果，赛后同步后自动回填。</td></tr>'
    return "".join(
        "<tr>"
        f"<td>{esc(item['date_text'])}</td>"
        f"<td>{esc(item['match_no'])}</td>"
        f"<td>{esc(item['home_team'])} vs {esc(item['away_team'])}</td>"
        f"<td><strong>{esc(item['score'])}</strong></td>"
        f"<td>{esc(item['result'])}</td>"
        f"<td>{esc(item['prediction_main_score'])}</td>"
        f"<td>{esc(item['review'])}</td>"
        "</tr>"
        for item in rows
    )


def prediction_match_lookup() -> dict[str, dict[str, Any]]:
    lookup: dict[str, dict[str, Any]] = {}
    for path in sorted(DATA_DIR.glob("knockout_predictions_*.json")):
        if re.search(r"_\d{4}$", path.stem):
            continue
        try:
            payload = read_json(path)
        except Exception:
            continue
        date_label = str(payload.get("date") or path.stem.rsplit("_", 1)[-1])
        for item in payload.get("matches", []):
            copied = dict(item)
            copied["_prediction_date"] = date_label
            match_no = str(item.get("match_no", "")).strip().lstrip("0")
            game_id = str(item.get("game_id", "")).strip()
            if match_no:
                lookup[f"no:{match_no}"] = copied
            if game_id:
                lookup[f"game:{game_id}"] = copied
    return lookup


def actual_half_direction(item: dict[str, Any]) -> str | None:
    for key in ("half_time_score", "half_time", "half_score"):
        value = item.get(key)
        if value:
            return half_direction_from_text(str(value))
    home = item.get("half_home_goals")
    away = item.get("half_away_goals")
    if home is None or away is None:
        return None
    try:
        home_goals = int(home)
        away_goals = int(away)
    except (TypeError, ValueError):
        return None
    if home_goals > away_goals:
        return "home"
    if home_goals < away_goals:
        return "away"
    return "draw"


def predicted_half_direction(item: dict[str, Any]) -> str | None:
    value = item.get("half_time_pick")
    if value in ("home", "draw", "away"):
        return str(value)
    value = item.get("half_time")
    if isinstance(value, dict):
        direction = value.get("direction")
        if direction in ("home", "draw", "away"):
            return str(direction)
        value = value.get("main_score") or value.get("score")
    if value:
        return half_direction_from_text(str(value))
    return None


def pct(hit: int, total: int) -> str:
    if total <= 0:
        return "暂无"
    return f"{hit}/{total} ({hit / total:.0%})"


def knockout_hit_stats() -> dict[str, Any]:
    predictions = prediction_match_lookup()
    rows = []
    score_hit = direction_hit = total_hit = half_hit = 0
    scored = direction_total = total_total = half_total = 0
    parlay_groups: dict[str, list[dict[str, Any]]] = {}

    for result in knockout_results():
        match_no = str(result.get("match_no", "")).strip().lstrip("0")
        pred = predictions.get(f"no:{match_no}")
        if not pred:
            continue
        actual_score = str(result.get("score", ""))
        predicted_score = str(pred.get("main_score", result.get("prediction_main_score", "")))
        actual_total = score_total(actual_score)
        predicted_total = score_total(predicted_score)
        actual_direction = score_direction(actual_score)
        predicted_direction = score_direction(predicted_score)
        actual_half = actual_half_direction(result)
        predicted_half = predicted_half_direction(pred)

        score_ok = predicted_score == actual_score
        direction_ok = predicted_direction == actual_direction
        total_ok = actual_total is not None and predicted_total is not None and actual_total == predicted_total
        half_ok = actual_half is not None and predicted_half is not None and actual_half == predicted_half

        scored += 1
        direction_total += 1
        total_total += 1
        score_hit += int(score_ok)
        direction_hit += int(direction_ok)
        total_hit += int(total_ok)
        if actual_half is not None and predicted_half is not None:
            half_total += 1
            half_hit += int(half_ok)

        row = {
            "date": result["date"],
            "match_no": match_no,
            "score_hit": score_ok,
            "direction_hit": direction_ok,
            "total_hit": total_ok,
            "half_hit": half_ok if actual_half is not None and predicted_half is not None else None,
        }
        rows.append(row)
        parlay_groups.setdefault(str(result["date"]), []).append(row)

    parlay_rows = []
    parlay_score_hit = parlay_total_hit = parlay_half_hit = 0
    parlay_score_total = parlay_total_total = parlay_half_total = 0
    for date_label, items in sorted(parlay_groups.items()):
        if len(items) != 3:
            continue
        score_ok = all(item["score_hit"] for item in items)
        total_ok = all(item["total_hit"] for item in items)
        half_available = all(item["half_hit"] is not None for item in items)
        half_ok = half_available and all(bool(item["half_hit"]) for item in items)
        parlay_score_total += 1
        parlay_total_total += 1
        parlay_score_hit += int(score_ok)
        parlay_total_hit += int(total_ok)
        if half_available:
            parlay_half_total += 1
            parlay_half_hit += int(half_ok)
        parlay_rows.append(
            {
                "date": date_label,
                "score": score_ok,
                "total": total_ok,
                "half": half_ok if half_available else None,
            }
        )

    return {
        "score": {"hit": score_hit, "total": scored},
        "direction": {"hit": direction_hit, "total": direction_total},
        "total_goals": {"hit": total_hit, "total": total_total},
        "half": {"hit": half_hit, "total": half_total},
        "parlay_score": {"hit": parlay_score_hit, "total": parlay_score_total},
        "parlay_total": {"hit": parlay_total_hit, "total": parlay_total_total},
        "parlay_half": {"hit": parlay_half_hit, "total": parlay_half_total},
        "parlay_rows": parlay_rows,
    }


def render_hit_metric(label: str, stat: dict[str, int]) -> str:
    return f'<div class="metric">{esc(label)}<strong>{esc(pct(stat["hit"], stat["total"]))}</strong><small>已结算样本</small></div>'


def render_parlay_hit_rows(rows: list[dict[str, Any]]) -> str:
    if not rows:
        return '<tr><td colspan="4">暂无满足 3 场的已结算日期。</td></tr>'
    return "".join(
        "<tr>"
        f"<td>{esc(row['date'][4:6] + '-' + row['date'][6:8])}</td>"
        f"<td>{'命中' if row['score'] else '未中'}</td>"
        f"<td>{'命中' if row['total'] else '未中'}</td>"
        f"<td>{'命中' if row['half'] is True else ('未中' if row['half'] is False else '半场赛果未回填')}</td>"
        "</tr>"
        for row in rows
    )


def render_knockout_hit_stats(stats: dict[str, Any]) -> str:
    return f"""
  <section class="section" id="hitStats">
    <h2>淘汰赛命中统计</h2>
    <div class="heroGrid">
      {render_hit_metric("精确比分", stats["score"])}
      {render_hit_metric("胜平负方向", stats["direction"])}
      {render_hit_metric("总进球数", stats["total_goals"])}
      {render_hit_metric("半场胜平负", stats["half"])}
      {render_hit_metric("比分三串一", stats["parlay_score"])}
      {render_hit_metric("总进球三串一", stats["parlay_total"])}
      {render_hit_metric("半场三串一", stats["parlay_half"])}
    </div>
    <div class="card tableWrap"><table><thead><tr><th>日期</th><th>比分三串一</th><th>总进球三串一</th><th>半场三串一</th></tr></thead><tbody>{render_parlay_hit_rows(stats["parlay_rows"])}</tbody></table></div>
  </section>"""


def future_text(matches: list[dict[str, Any]]) -> str:
    return "\n".join(
        f"{item['beijing_date'][5:]} {item['beijing_time']}  {item['match_no']}  {item['home_team']} vs {item['away_team']}  {item['round']}"
        for item in matches
    )


def date_dirs() -> list[Path]:
    return sorted(
        [p for p in ROOT.glob("2026*") if p.is_dir() and re.match(r"^2026\d{4}$", p.name)],
        key=lambda p: p.name,
        reverse=True,
    )


def prediction_payload_paths() -> list[Path]:
    paths = []
    for path in sorted(DATA_DIR.glob("knockout_predictions_*.json")):
        date_label = path.stem.rsplit("_", 1)[-1]
        if re.match(r"^2026\d{4}$", date_label):
            paths.append(path)
    return paths


def read_prediction_payloads() -> dict[str, dict[str, Any]]:
    payloads: dict[str, dict[str, Any]] = {}
    for path in prediction_payload_paths():
        try:
            payload = read_json(path)
        except Exception:
            continue
        date_label = str(payload.get("date") or path.stem.rsplit("_", 1)[-1])
        if re.match(r"^2026\d{4}$", date_label):
            payloads[date_label] = payload
    return payloads


def knockout_archive_label(date_label: str, payload: dict[str, Any] | None = None) -> str:
    labels = {
        "20260629": "南非 vs 加拿大",
    }
    if date_label in labels:
        return labels[date_label]
    matches = (payload or {}).get("matches", [])
    teams = [str(item.get("home_team", "")).strip() for item in matches if item.get("home_team")]
    if len(teams) >= 3:
        return "/".join(teams[:3]) + "三场"
    if len(matches) == 1:
        item = matches[0]
        return f"{item.get('home_team', '')} vs {item.get('away_team', '')}"
    if matches:
        return f"{len(matches)}场淘汰赛预测"
    return "淘汰赛预测"


def render_date_cards(prefix: str = "") -> str:
    cards = []
    for p in date_dirs():
        stage = "淘汰赛" if p.name >= "20260629" else "小组赛"
        href = f"{prefix}{p.name}/"
        subtitle = "南非 vs 加拿大" if p.name == "20260629" else "历史预测与复盘"
        cards.append(
            f'<a class="dateCard" href="{href}"><span>{p.name}</span><strong>{stage}</strong><em>{subtitle}</em></a>'
        )
    return "".join(cards)


def render_group_cards(prefix: str = "") -> str:
    cards = []
    for p in date_dirs():
        if p.name >= "20260629":
            continue
        href = f"{prefix}{p.name}/"
        cards.append(f'<a class="miniCard" href="{href}"><strong>{p.name}</strong><span>小组赛预测</span></a>')
    return "".join(cards)


def home_stats_subset(stats: list[dict[str, Any]], matches: list[dict[str, Any]]) -> list[dict[str, Any]]:
    future_names = {m["home_team"] for m in matches} | {m["away_team"] for m in matches}
    top_by_wins = sorted(stats, key=lambda x: (-x["wins"], -x["points"], -x["goal_difference"], x["team_name"]))[:3]
    selected = []
    seen = set()
    for item in top_by_wins + [s for s in stats if s["team_name"] in future_names]:
        if item["team_name"] not in seen:
            selected.append(item)
            seen.add(item["team_name"])
    return selected


def score_direction(score: str) -> str:
    try:
        left, right = re.split(r"[:：-]", score)[:2]
        home, away = int(left), int(right)
    except Exception:
        return "unknown"
    if home > away:
        return "home"
    if home < away:
        return "away"
    return "draw"


def model_review_lessons() -> dict[str, Any]:
    settled = []
    for path in sorted(DATA_DIR.glob("202606*.json")):
        try:
            payload = read_json(path)
        except Exception:
            continue
        for match in payload.get("matches", []):
            result = match.get("result")
            prediction = match.get("prediction", {})
            if not result or result.get("status") != "Finished":
                continue
            home_goals = result.get("homeGoals")
            away_goals = result.get("awayGoals")
            if home_goals is None or away_goals is None:
                continue
            actual_score = f"{home_goals}:{away_goals}"
            predicted_scores = prediction.get("scores", [])
            predicted_total = str(prediction.get("totalGoals", ""))
            first_score = predicted_scores[0] if predicted_scores else ""
            settled.append(
                {
                    "date": path.stem,
                    "match": f"{match.get('home')} vs {match.get('away')}",
                    "predicted_total": predicted_total,
                    "predicted_scores": predicted_scores,
                    "actual_score": actual_score,
                    "actual_total": int(home_goals) + int(away_goals),
                    "direction_hit": score_direction(first_score) == score_direction(actual_score),
                    "score_hit": actual_score in predicted_scores,
                    "total_hit": predicted_total == str(int(home_goals) + int(away_goals)),
                }
            )
    total = len(settled)
    if not total:
        return {"settled": 0, "lessons": ["本地暂未找到可结算赛果。"], "summary": "暂无足够复盘样本。"}
    direction_hits = sum(1 for row in settled if row["direction_hit"])
    total_hits = sum(1 for row in settled if row["total_hit"])
    score_hits = sum(1 for row in settled if row["score_hit"])
    high_miss = sum(1 for row in settled if row["actual_total"] >= 4 and row["predicted_total"] in {"1", "2"})
    low_miss = sum(1 for row in settled if row["actual_total"] <= 1 and row["predicted_total"] in {"3", "4", "5"})
    lessons = [
        f"已回填 {total} 场小组赛赛果：方向比比分更可靠，比分不能只盯一个主比分。",
        f"方向命中约 {direction_hits}/{total}，说明强弱判断可保留；总进球命中约 {total_hits}/{total}，说明进球区间要用范围表达。",
        f"出现 {high_miss} 场低估大比分、{low_miss} 场高估进球的偏差；淘汰赛模型必须同时保留低比分锁局和尾部上修。",
        f"精确比分命中 {score_hits}/{total}，所以今天输出改为“主比分 + 两条备选脚本”，不再把单一比分说死。",
        "南非-加拿大 0-1 复盘：模型低比分和加拿大最终优势方向有效，但90分钟平局保护偏重；补时进球说明强侧尾段小胜权重需要上调。",
        "6月30日三场复盘：巴西-日本 2-1 命中主比分，荷兰-摩洛哥 1-1 命中主比分，德国-巴拉圭 1-1 命中冷门保护但主线过强。",
        "今日修正：强方小胜保留，但面对低位防守、定位球强队和窄差对抗时，上调90分钟平局、1-1、0-0/1-0低比分权重。",
        "末轮阶段常被净胜球需求放大，淘汰赛首战则相反：先压节奏，再看60分钟后的体能、换人和补时阶段。",
    ]
    return {
        "settled": total,
        "direction_hits": direction_hits,
        "total_hits": total_hits,
        "score_hits": score_hits,
        "high_miss": high_miss,
        "low_miss": low_miss,
        "lessons": lessons,
        "headline": "6月30日复盘：巴西2-1、日本进球、荷兰1-1命中；德国被巴拉圭拖成1-1，提醒模型继续上调平局和定位球冷门保护。",
        "summary": "复盘后模型口径：强方一球小胜仍可作为主线，但淘汰赛窄差局优先保留1-1、低比分和最终晋级/90分钟赛果拆分。",
    }


def css() -> str:
    return """
:root{--bg:#06120f;--panel:#102528;--soft:#0b201d;--line:#1f4a43;--text:#e9fff8;--muted:#9bb8b0;--green:#33e28a;--blue:#7dd3fc;--orange:#ffad4d;--red:#ff5a63}
*{box-sizing:border-box}body{margin:0;font-family:"Microsoft YaHei",Arial,sans-serif;background:#05110f;color:var(--text)}a{color:inherit;text-decoration:none}
body:before{content:"";position:fixed;inset:0;background:linear-gradient(135deg,#04120f 0%,#092225 48%,#061324 100%);z-index:-2}
body:after{content:"";position:fixed;inset:0;background:radial-gradient(circle at 50% -20%,rgba(51,226,138,.13),transparent 42%);z-index:-1;pointer-events:none}
header,main{max-width:1180px;margin:auto;padding:24px 18px}header{position:sticky;top:0;background:rgba(3,12,11,.92);backdrop-filter:blur(10px);border-bottom:1px solid var(--line);z-index:5}
h1{margin:0 0 10px;font-size:clamp(28px,5vw,42px);letter-spacing:0}.topbar{display:flex;justify-content:space-between;gap:14px;align-items:flex-start}.topActions{display:flex;gap:10px;flex-wrap:wrap;justify-content:flex-end}@media(max-width:850px){.topbar{display:block}.topActions{justify-content:flex-start;margin-top:12px}}
nav{display:flex;gap:10px;flex-wrap:wrap;margin-top:14px}nav a,.buttonLink{padding:9px 12px;border:1px solid var(--line);border-radius:8px;background:#0b201d;color:#8fffd0;display:inline-flex;align-items:center;gap:6px}
.section{margin:24px 0 36px}.section h2{color:var(--blue);font-size:25px;margin:0 0 12px}.card,.panel{border:1px solid var(--line);border-radius:8px;background:rgba(12,32,33,.96);box-shadow:0 16px 36px rgba(0,0,0,.22);padding:18px;margin:14px 0}
.entryGrid{display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:18px}.entryCard{min-height:170px;border:1px solid var(--line);border-radius:8px;background:linear-gradient(180deg,#0f2a2b,#091b1d);padding:22px;display:flex;flex-direction:column;justify-content:space-between}.entryCard:hover{border-color:var(--green)}.entryCard span{color:var(--muted)}.entryCard strong{font-size:30px;color:var(--green);display:block;margin:8px 0}.entryCard em{font-style:normal;color:var(--text);line-height:1.6}
.heroGrid,.coreGrid,.summaryGrid{display:grid;grid-template-columns:repeat(4,minmax(0,1fr));gap:12px}.coreGrid{grid-template-columns:repeat(3,minmax(0,1fr))}.summaryGrid{grid-template-columns:repeat(3,minmax(0,1fr))}@media(max-width:850px){.entryGrid,.heroGrid,.coreGrid,.summaryGrid{grid-template-columns:1fr}}
.metric{background:#071817;border:1px solid #1f4a43;border-radius:8px;padding:14px}.metric strong{display:block;color:var(--green);font-size:23px;margin-top:4px}.metric small{display:block;color:var(--muted);margin-top:5px;line-height:1.5}
.bigScore{font-size:48px;color:var(--green);font-weight:900}.adv{color:var(--orange);font-weight:900}.muted{color:var(--muted)}.tableWrap{overflow-x:auto}table{width:100%;border-collapse:collapse;min-width:820px}th,td{padding:11px;border-bottom:1px solid var(--line);text-align:left;vertical-align:top}th{color:#8fffd0;background:#09211e}
ul{line-height:1.8}.scoreTag{display:inline-block;margin:4px 8px 4px 0;padding:8px 12px;border-radius:8px;border:1px solid #2ecf7a;background:#0d3028;color:#a8ffd6;font-weight:800}.danger{border-color:var(--red);color:#ffd2b0;background:#331415}
.resultTable td{padding:14px 12px;line-height:1.65}.resultTable td:last-child{min-width:360px}.resultTable strong{font-size:18px;color:#a8ffd6}
.miniGrid{display:grid;grid-template-columns:repeat(auto-fill,minmax(150px,1fr));gap:12px}.miniCard{display:flex;flex-direction:column;gap:7px;padding:14px;border:1px solid var(--line);border-radius:8px;background:#071817}.miniCard:hover{border-color:var(--green)}.miniCard span{color:var(--muted)}
.mysticGrid{display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:12px}@media(max-width:850px){.mysticGrid{grid-template-columns:1fr}}.mysticItem{border:1px solid var(--line);border-radius:8px;background:#071817;padding:14px}.mysticItem strong{color:#8fffd0;display:block;margin-bottom:6px}
.mysticCard{border-color:#7c3aed;background:linear-gradient(180deg,rgba(43,25,74,.96),rgba(20,13,38,.96))}.mysticCard .mysticItem{border-color:#6d46b4;background:rgba(26,16,48,.9)}.mysticCard .mysticItem strong{color:#d8b4fe}.mysticSummary{display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:12px;margin-top:14px}@media(max-width:850px){.mysticSummary{grid-template-columns:1fr}}.mysticSummary div{border:1px solid #6d46b4;border-radius:8px;background:rgba(26,16,48,.9);padding:12px}.mysticSummary strong{display:block;color:#f0abfc;margin-bottom:5px}
.matchDetails{border:1px solid var(--line);border-radius:8px;background:rgba(7,24,23,.96);overflow:hidden}.matchDetails summary{cursor:pointer;list-style:none;display:grid;grid-template-columns:minmax(180px,1fr) 90px minmax(240px,1fr);gap:12px;align-items:center;padding:16px 18px;border-bottom:1px solid var(--line)}.matchDetails summary::-webkit-details-marker{display:none}.matchDetails summary:before{content:"展开";color:#8fffd0;border:1px solid var(--line);border-radius:8px;padding:5px 8px;justify-self:start;grid-column:1/-1;display:none}.matchDetails:not([open]) summary:before{display:inline-block}.matchDetails summary span{font-size:22px;font-weight:800;color:var(--blue)}.matchDetails summary strong{font-size:30px;color:var(--green)}.matchDetails summary em{font-style:normal;color:var(--muted);line-height:1.5}.matchDetails:not([open]) summary{border-bottom:0}.matchBody{border:0;border-top:1px solid var(--line);border-radius:0;margin:0;box-shadow:none}@media(max-width:850px){.matchDetails summary{grid-template-columns:1fr}.matchDetails summary strong{font-size:28px}}
footer{color:#8ea8a1;font-size:13px;margin-top:36px}
"""


def render_home_page(schedule: list[dict[str, Any]], stats: list[dict[str, Any]], prediction: dict[str, Any]) -> str:
    matches = rolling_future_matches(schedule, 3)
    round16_matches = future_round16_matches(schedule)
    round16_forecasts = round16_market_forecasts()
    quarterfinals = quarterfinal_matches(schedule)
    results = knockout_results()
    hit_stats = knockout_hit_stats()
    return f"""<!DOCTYPE html>
<html lang="zh-CN">
<head><meta charset="UTF-8"><meta name="viewport" content="width=device-width, initial-scale=1.0"><title>2026世界杯预测入口</title><style>{css()}</style></head>
<body>
<header>
  <div class="topbar"><div><h1>2026世界杯预测入口</h1>
  </div></div>
  <nav><a href="#quarterfinals">8强入围</a><a href="#future">未来三天</a><a href="#round16">16强赛程</a><a href="#round16Goals">16强总进球</a><a href="parlay/">未来串关</a><a href="#hitStats">命中统计</a><a href="#teams">淘汰赛赛果</a></nav>
</header>
<main>
  <section class="section" id="quarterfinals"><h2>8强比赛入围表</h2><div class="card tableWrap"><table><thead><tr><th>北京时间</th><th>场次</th><th>上半区入围队</th><th>状态</th><th>下半区入围队</th><th>状态</th></tr></thead><tbody>{render_quarterfinal_entry_rows(quarterfinals)}</tbody></table></div></section>
  <section class="section" id="entry">
    <h2>赛事阶段</h2>
    <div class="entryGrid">
      <a class="entryCard" href="group/"><span>历史归档</span><strong>小组赛</strong><em>按日期查看 20260617-20260628 的历史预测和复盘。</em></a>
      <a class="entryCard" href="knockout/"><span>当前阶段</span><strong>淘汰赛</strong><em>按日期查看淘汰赛预测。</em></a>
    </div>
  </section>
  <section class="section" id="future"><h2>未来三天比赛</h2><div class="card tableWrap"><table><thead><tr><th>北京时间</th><th>场次</th><th>轮次</th><th>对阵</th><th>下场对手</th></tr></thead><tbody>{render_upcoming_rows(matches)}</tbody></table></div></section>
  <section class="section" id="round16"><h2>未来16强赛日程</h2><div class="card tableWrap"><table><thead><tr><th>北京时间</th><th>场次</th><th>对阵</th><th>晋级路径</th></tr></thead><tbody>{render_round16_schedule_rows(round16_matches)}</tbody></table></div></section>
  <section class="section" id="round16Goals"><h2>未来16强总进球预测</h2><div class="card tableWrap"><table><thead><tr><th>场次</th><th>北京时间</th><th>对阵</th><th>主推总进球</th><th>核心区间</th><th>备选</th><th>模型依据</th></tr></thead><tbody>{render_round16_total_goals_rows(round16_forecasts)}</tbody></table></div></section>
  {render_knockout_hit_stats(hit_stats)}
  <section class="section" id="teams"><h2>淘汰赛以来赛果</h2><div class="card tableWrap"><table class="resultTable"><thead><tr><th>日期</th><th>场次</th><th>对阵</th><th>比分</th><th>赛果</th><th>赛前主比分</th><th>复盘</th></tr></thead><tbody>{render_knockout_result_rows(results)}</tbody></table></div></section>
</main>
</body>
</html>"""


def render_mystic_framework(mystic: dict[str, Any]) -> str:
    data_pack = "".join(
        f"<div class=\"mysticItem\"><strong>{esc(item['field'])}</strong><p>{esc(item['value'])}</p></div>"
        for item in mystic["data_pack"]
    )
    items = "".join(
        f"<div class=\"mysticItem\"><strong>{esc(item['name'])}</strong><p>{esc(item['content'])}</p></div>"
        for item in mystic["framework"]
    )
    summary = "".join(
        f"<div><strong>{esc(key)}</strong><p>{esc(value)}</p></div>"
        for key, value in mystic["summary"].items()
    )
    return (
        f"<p>{esc(mystic['status'])}</p>"
        f"<p>{esc(mystic['local_time_rule'])}</p>"
        f"<h3>基础包</h3><div class=\"mysticGrid\">{data_pack}</div>"
        f"<h3>术数拆解</h3>"
        f"<div class=\"mysticGrid\">{items}</div>"
        f"<h3>综合结论</h3><div class=\"mysticSummary\">{summary}</div>"
        f"<p class=\"muted\">{esc(mystic['reading'])}</p>"
    )


def render_teams_page(stats: list[dict[str, Any]], schedule: list[dict[str, Any]]) -> str:
    rows = render_stats_rows(stats)
    upcoming = render_upcoming_rows(rolling_future_matches(schedule, 3))
    return f"""<!DOCTYPE html>
<html lang="zh-CN">
<head><meta charset="UTF-8"><meta name="viewport" content="width=device-width, initial-scale=1.0"><title>32强当前战绩</title><style>{css()}</style></head>
<body>
<header>
  <h1>32强当前战绩</h1>
  <nav><a href="../index.html">首页</a><a href="../20260701/">淘汰赛预测</a><a href="#teams">完整战绩</a><a href="#future">未来三天</a></nav>
</header>
<main>
  <section class="section" id="teams"><h2>完整 32 强战绩</h2><div class="card tableWrap"><table><thead><tr><th>球队</th><th>代码</th><th>小组</th><th>出线路径</th><th>胜</th><th>平</th><th>负</th><th>进</th><th>失</th><th>净胜</th><th>积分</th></tr></thead><tbody>{rows}</tbody></table></div></section>
  <section class="section" id="future"><h2>未来三天赛程</h2><div class="card tableWrap"><table><thead><tr><th>北京时间</th><th>场次</th><th>轮次</th><th>对阵</th><th>下场对手</th></tr></thead><tbody>{upcoming}</tbody></table></div></section>
</main>
</body>
</html>"""


def render_group_archive_page() -> str:
    cards = render_group_cards("../")
    return f"""<!DOCTYPE html>
<html lang="zh-CN">
<head><meta charset="UTF-8"><meta name="viewport" content="width=device-width, initial-scale=1.0"><title>小组赛预测归档</title><style>{css()}</style></head>
<body>
<header>
  <h1>小组赛预测归档</h1>
  <nav><a href="../index.html">首页</a><a href="../20260701/">淘汰赛预测</a></nav>
</header>
<main>
  <section class="section"><h2>按日期查看</h2><div class="miniGrid">{cards}</div></section>
</main>
</body>
</html>"""


def render_knockout_archive_page() -> str:
    payloads = read_prediction_payloads()
    cards = "".join(
        f'<a class="miniCard" href="../{p.name}/"><strong>{p.name}</strong><span>{knockout_archive_label(p.name, payloads.get(p.name))}</span></a>'
        for p in date_dirs()
        if p.name >= "20260629"
    )
    return f"""<!DOCTYPE html>
<html lang="zh-CN">
<head><meta charset="UTF-8"><meta name="viewport" content="width=device-width, initial-scale=1.0"><title>淘汰赛预测归档</title><style>{css()}</style></head>
<body>
<header>
  <h1>淘汰赛预测归档</h1>
  <nav><a href="../index.html">首页</a></nav>
</header>
<main>
  <section class="section"><h2>按日期查看</h2><div class="miniGrid">{cards}</div></section>
</main>
</body>
</html>"""


def render_prediction_card(item: dict[str, Any]) -> str:
    team_reading = render_bullets(item.get("team_reading", []))
    model_reasoning = render_bullets(item.get("model_reasoning", item.get("triggers", [])))
    mystic = render_bullets(item.get("mystic_summary", ["仅作娱乐表达，不进入模型权重。"]))
    triggers = render_bullets(item.get("triggers", []))
    matrix_rows = render_score_matrix(item.get("score_matrix_top", []))
    ev_rows = render_ev_rows(item.get("ev_table", []))
    insight_cards = render_insight_cards(item.get("model_insights", []))
    compact_rows = compact_score_rows(item.get("score_matrix_top", []))
    safe_scores = [str(row.get("score")) for row in compact_rows if "冷" not in str(row.get("ev_signal", "")) and "冷" not in str(row.get("label", ""))]
    backup = " / ".join(safe_scores[1:3] or item.get("backup_scores", item.get("safe_scores", []))[:2])
    score_tags = safe_scores[:3]
    if not score_tags:
        score_tags = [item.get("main_score")]
    score_tag_html = "".join(f'<span class="scoreTag">{esc(score)}</span>' for score in score_tags if score)
    if item.get("upset_score"):
        score_tag_html += f'<span class="scoreTag danger">{esc(item["upset_score"])}</span>'
    kickoff = str(item.get("kickoff_bjt", "")).replace(" +08:00", "")
    half_time = item.get("half_time")
    if not half_time:
        half_time = half_result_label(str(item.get("half_time_pick", "draw")))
    next_note = item.get("next_opponent_note", "下场对手待定")
    return f"""
<section class="section">
  <details class="matchDetails">
  <summary>
    <span>{esc(item['home_team'])} vs {esc(item['away_team'])}</span>
    <strong>{esc(item['main_score'])}</strong>
    <em>{esc(item['direction_text'])} / 晋级 {esc(item['advance_pick'])}</em>
  </summary>
  <div class="card matchBody">
    <div class="heroGrid">
      <div class="metric">北京时间<strong>{esc(kickoff)}</strong></div>
      <div class="metric">半场<strong>{esc(half_time)}</strong></div>
      <div class="metric">90分钟<strong>{esc(item['direction_text'])}</strong></div>
      <div class="metric">晋级<strong class="adv">{esc(item['advance_pick'])}</strong></div>
    </div>
    <div class="coreGrid">
      <div class="panel"><h3>主比分</h3><div class="bigScore">{esc(item['main_score'])}</div><p>备选：{esc(backup)}</p><p>冷门：{esc(item['upset_score'])}</p></div>
      <div class="panel"><h3>判断口径</h3><p>{esc(item['core_view'])}</p><p>{esc(item['advance_probability_text'])}</p></div>
      <div class="panel"><h3>下场对手</h3><p>{esc(next_note)}</p></div>
    </div>
    <div class="card">
      <h3>昨日复盘校正</h3>
      <p>{esc(item.get('review_adjustment', '南非-加拿大0-1复盘后，今日统一提高低比分、尾段小胜和触发条件权重。'))}</p>
    </div>
    <div class="card">
      <h3>比分池</h3>
      {score_tag_html}
      <div class="tableWrap"><table><thead><tr><th>比分</th><th>概率</th><th>EV信号</th><th>标签</th></tr></thead><tbody>{matrix_rows}</tbody></table></div>
    </div>
    <div class="card">
      <h3>市场 EV / 半场 / 胜负</h3>
      <div class="tableWrap"><table><thead><tr><th>市场</th><th>模型概率</th><th>公平赔率</th><th>EV信号</th><th>说明</th></tr></thead><tbody>{ev_rows}</tbody></table></div>
    </div>
    <div class="coreGrid">
      {insight_cards}
    </div>
    <div class="card"><h3>触发条件</h3><ul>{triggers}</ul></div>
    <div class="coreGrid">
      <div class="panel"><h3>球队画像</h3><ul>{team_reading}</ul></div>
      <div class="panel"><h3>模型解释</h3><ul>{model_reasoning}</ul></div>
      <div class="panel mysticCard"><h3>玄学综参</h3><ul>{mystic}</ul></div>
    </div>
  </div>
  </details>
</section>"""


def render_parlay_table(title: str, rows: list[dict[str, Any]], note: str) -> str:
    body = "".join(
        "<tr>"
        f"<td>{esc(row['match'])}</td><td>{esc(row['pick'])}</td><td>{esc(row.get('odds_text', fmt_odds(row.get('odds'))))}</td>"
        f"<td>{esc(row.get('updated_at', '-'))}</td></tr>"
        for row in rows
    )
    total = product_odds(rows)
    return f"""
    <div class="card">
      <h3>{esc(title)}</h3>
      <p class="muted">{esc(note)}</p>
      <div class="tableWrap"><table><thead><tr><th>比赛</th><th>选择</th><th>赔率</th><th>赔率更新时间</th></tr></thead><tbody>{body}</tbody></table></div>
      <p><strong>三串一理论乘积赔率：{esc(fmt_odds(total))}</strong></p>
    </div>"""


def render_parlay_section(date_label: str, predictions: list[dict[str, Any]]) -> str:
    settled = settled_match_keys()
    predictions = [
        item
        for item in predictions
        if f"game:{item.get('game_id', '')}" not in settled
        and f"no:{str(item.get('match_no', '')).lstrip('0')}" not in settled
    ]
    if len(predictions) < PARLAY_POLICY["min_enabled_match_count"]:
        return ""
    odds_lookup = load_live_odds(date_label)

    score_rows: list[dict[str, Any]] = []
    goals_rows: list[dict[str, Any]] = []
    half_rows: list[dict[str, Any]] = []

    for item in predictions:
        odds_match = odds_match_for_prediction(item, odds_lookup) if odds_lookup else None
        odds = odds_match.get("odds", {}) if odds_match else {}
        match_label = f"{item['home_team']} vs {item['away_team']}"

        positive_scores = [
            str(row.get("score"))
            for row in item.get("score_matrix_top", [])
            if str(row.get("ev_signal", "")).startswith("+")
        ][:2]
        if not positive_scores:
            positive_scores = [str(item["main_score"])]
        crs = odds.get("crs") or {}
        score_odds = [fmt_odds(odds_float(crs.get(score))) for score in positive_scores]
        score_rows.append(
            {
                "match": match_label,
                "pick": " / ".join(positive_scores),
                "odds": None,
                "odds_text": " / ".join(score_odds) if score_odds else "-",
                "updated_at": crs.get("updatedAt", "-"),
            }
        )

        ttg = odds.get("ttg") or {}
        total_picks = []
        for score in positive_scores:
            total = score_total(score)
            if total is not None and total not in total_picks:
                total_picks.append(total)
        total_odds = []
        for total in total_picks:
            ttg_key = "s7" if total >= 7 else f"s{total}"
            total_odds.append(fmt_odds(odds_float(ttg.get(ttg_key))))
        goals_rows.append(
            {
                "match": match_label,
                "pick": " / ".join(f"总进球 {total}" for total in total_picks) if total_picks else "总进球 -",
                "odds": None,
                "odds_text": " / ".join(total_odds) if total_odds else "-",
                "updated_at": ttg.get("updatedAt", "-"),
            }
        )

        half_value = item.get("half_time")
        if half_value:
            half_direction = half_direction_from_text(str(half_value))
        else:
            half_direction = str(item.get("half_time_pick", "draw"))
        full_direction = str(item.get("direction_90min") or score_direction(str(item.get("main_score", ""))))
        hafu_pick_key = hafu_key_from_label_or_key(str(item.get("hafu_pick", ""))) if item.get("hafu_pick") else hafu_key(half_direction, full_direction)
        hafu = odds.get("hafu") or {}
        half_rows.append(
            {
                "match": match_label,
                "pick": hafu_label_from_key(hafu_pick_key),
                "odds": odds_float(hafu.get(hafu_pick_key)),
                "updated_at": hafu.get("updatedAt", "-"),
            }
        )

    score_html = render_parlay_table("比分三串一", score_rows, "按每场保守EV正信号的两个比分选择，赔率来自体彩比分市场；缺少赔率快照时显示 -。")
    goals_html = render_parlay_table("总进球数三串一", goals_rows, "按每场两个保守EV正信号比分折算总进球数，赔率来自体彩总进球市场；缺少赔率快照时显示 -。")
    half_html = render_parlay_table("半全场胜平负三串一", half_rows, "按每场半全场组合选择，赔率来自体彩半全场 hafu 赔率池。")
    return f"""
  <section class="section" id="parlay">
    <h2>今日三串一</h2>
    {score_html}
    {goals_html}
    {half_html}
  </section>"""


def previous_date_label(date_label: str) -> str:
    try:
        return (datetime.strptime(date_label, "%Y%m%d") - timedelta(days=1)).strftime("%Y%m%d")
    except ValueError:
        return date_label


def model_review_for_prediction_date(date_label: str) -> dict[str, Any]:
    candidates = [
        DATA_DIR / f"model_review_lessons_{previous_date_label(date_label)}.json",
        DATA_DIR / f"model_review_lessons_{date_label}.json",
    ]
    for path in candidates:
        if not path.exists():
            continue
        try:
            return read_json(path)
        except Exception:
            continue
    return model_review_lessons()


def render_extra_review_section(review: dict[str, Any]) -> str:
    text = " ".join(str(item) for item in review.get("lessons", [])) + " " + str(review.get("summary", ""))
    if "瑞士" not in text or "2-0" not in text:
        return ""
    return """
  <section class="section" id="swiss">
    <h2>瑞士 2-0 漏出原因</h2>
    <div class="card">
      <p>瑞士场不是方向问题，而是比分池结构问题：模型把 <strong>1-1最低比分赔率</strong> 和淘汰赛防平权重放得过高，却没有把 <strong>主胜、总进球2球最低、对手终结弱、瑞士低风险零封</strong> 合并成 2-0 路径。修正后，中等优势队的低比分池必须同时检查 1-0、2-0、2-1、1-1。</p>
    </div>
  </section>"""


def render_multi_prediction_page(date_label: str, predictions: list[dict[str, Any]]) -> str:
    cards = "".join(render_prediction_card(item) for item in predictions)
    review = model_review_for_prediction_date(date_label)
    review_lessons = render_review_lessons(review)
    extra_review = render_extra_review_section(review)
    parlay = render_parlay_section(date_label, predictions)
    parlay_nav = '<a href="#parlay">三串一</a>' if parlay else ""
    return f"""<!DOCTYPE html>
<html lang="zh-CN">
<head><meta charset="UTF-8"><meta name="viewport" content="width=device-width, initial-scale=1.0"><title>{date_label} 淘汰赛预测</title><style>{css()}</style></head>
<body>
<header>
  <h1>{date_label} 淘汰赛预测</h1>
  <nav><a href="../index.html">首页</a><a href="../knockout/">淘汰赛日期</a>{parlay_nav}</nav>
</header>
<main>
  <section class="section">
    <h2>昨日复盘与模型修正</h2>
    <div class="card">
      <p><strong>{esc(review.get('headline', '昨日赛果已完成复盘并写入模型校正。'))}</strong></p>
      <p>{esc(review.get('summary', ''))}</p>
      <ul>{review_lessons}</ul>
    </div>
  </section>
  {extra_review}
  {parlay}
  {cards}
</main>
</body>
</html>"""


def render_prediction_page(schedule: list[dict[str, Any]], stats: list[dict[str, Any]], prediction: dict[str, Any]) -> str:
    mystic = mystic_panel()
    matrix_rows = render_score_matrix(prediction["score_matrix_top"])
    ev_rows = render_ev_rows(prediction["ev_table"])
    analysis = render_bullets(prediction["analysis"])
    risks = render_bullets(prediction["risks"])
    triggers = render_bullets(prediction["triggers"])
    insight_cards = render_insight_cards(prediction["model_insights"])
    plan_rows = render_match_plan(prediction["match_plan"])
    mystic_html = render_mystic_framework(mystic)
    next_opponent = prediction["scenario"]["next_opponent_source"]
    home_profile = render_bullets(prediction["team_profiles"]["home"])
    away_profile = render_bullets(prediction["team_profiles"]["away"])
    return f"""<!DOCTYPE html>
<html lang="zh-CN">
<head><meta charset="UTF-8"><meta name="viewport" content="width=device-width, initial-scale=1.0"><title>20260629 南非 vs 加拿大预测</title><style>{css()}</style></head>
<body>
<header>
  <h1>南非 vs 加拿大</h1>
  <nav><a href="../index.html">首页</a><a href="#summary">结论</a><a href="#profiles">球队画像</a><a href="#model">模型解释</a><a href="#mystic">玄学分析</a></nav>
</header>
<main>
  <section class="section" id="summary">
    <h2>核心结论</h2>
    <div class="card">
      <div class="heroGrid">
        <div class="metric">北京时间<strong>{esc(prediction['kickoff_bjt'].replace(' +08:00',''))}</strong><small>当地时间：{esc(prediction['kickoff_local'])}</small></div>
        <div class="metric">半场<strong>{esc(prediction['half_time']['main_score'])}</strong><small>方向：{esc(prediction['half_time']['direction'])}，备选 {esc(prediction['half_time']['backup_score'])}</small></div>
        <div class="metric">90分钟<strong>{esc(prediction['main_score'])}</strong><small>{esc(prediction['direction_text'])}，总进球 {esc(prediction['goals_range'])}</small></div>
        <div class="metric">晋级<strong class="adv">{esc(prediction['advance_pick'])}</strong><small>{esc(prediction['extra_time_penalty_script'])}</small></div>
      </div>
      <div class="coreGrid">
        <div class="panel"><h3>90分钟胜平负</h3><p>主胜 {prediction['full_time_90']['probabilities']['home']:.0%} / 平 {prediction['full_time_90']['probabilities']['draw']:.0%} / 客胜 {prediction['full_time_90']['probabilities']['away']:.0%}</p><p>{esc(prediction['full_time_90']['double_chance'])}</p></div>
        <div class="panel"><h3>晋级概率</h3><p>{esc(prediction['home_team'])}: {prediction['home_advance_probability']:.0%}</p><p>{esc(prediction['away_team'])}: {prediction['away_advance_probability']:.0%}</p></div>
        <div class="panel"><h3>下一场潜在对手</h3><p>{esc(next_opponent)}</p></div>
      </div>
    </div>
  </section>
  <section class="section" id="scores"><h2>比分与 EV</h2><div class="card"><h3>比分池</h3><span class="scoreTag">{esc(prediction['main_score'])}</span><span class="scoreTag">{esc(prediction['safe_scores'][0])}</span><span class="scoreTag">{esc(prediction['safe_scores'][1])}</span><span class="scoreTag">{esc(prediction['draw_protection_scores'][0])}</span><span class="scoreTag danger">{esc(prediction['upset_score'])}</span><div class="tableWrap"><table><thead><tr><th>比分</th><th>概率</th><th>EV信号</th><th>标签</th></tr></thead><tbody>{matrix_rows}</tbody></table></div></div><div class="card"><h3>市场 EV / 半场 / 胜负</h3><div class="tableWrap"><table><thead><tr><th>市场</th><th>模型概率</th><th>公平赔率</th><th>EV信号</th><th>说明</th></tr></thead><tbody>{ev_rows}</tbody></table></div></div></section>
  <section class="section" id="profiles"><h2>球队画像</h2><div class="coreGrid"><div class="panel"><h3>{esc(prediction['home_team'])}</h3><ul>{home_profile}</ul></div><div class="panel"><h3>{esc(prediction['away_team'])}</h3><ul>{away_profile}</ul></div><div class="panel"><h3>实力判断</h3><p>加拿大整体进攻更好，南非防守韧性更强。</p><p>所以本场不走“加拿大轻松赢”的路线，而是走“南非拖住90分钟，加拿大靠后程和加时/点球晋级”的路线。</p><p>模型结论：半场平，90分钟平局优先，最终晋级加拿大。</p></div></div><div class="coreGrid"><div class="panel"><h3>本场合理打法</h3><p><strong>{esc(prediction['game_plan']['home'])}</strong></p></div><div class="panel"><h3>本场合理打法</h3><p><strong>{esc(prediction['game_plan']['away'])}</strong></p></div><div class="panel"><h3>下一场潜在对手</h3><p>荷兰 / 摩洛哥胜者</p></div></div></section>
  <section class="section" id="model"><h2>模型解释</h2><div class="coreGrid">{insight_cards}</div><div class="card"><h3>比赛阶段脚本</h3><div class="tableWrap"><table><thead><tr><th>阶段</th><th>{esc(prediction['home_team'])}</th><th>{esc(prediction['away_team'])}</th></tr></thead><tbody>{plan_rows}</tbody></table></div></div><div class="card"><h3>触发条件</h3><ul>{triggers}</ul></div></section>
  <section class="section" id="story"><h2>比赛剧本与风险</h2><div class="coreGrid"><div class="panel"><h3>晋级剧本</h3><p>{esc(prediction['scenario']['home_win_script'])}</p><p>{esc(prediction['scenario']['away_win_script'])}</p></div><div class="panel"><h3>分析依据</h3><ul>{analysis}</ul></div><div class="panel"><h3>风险提示</h3><ul>{risks}</ul></div></div></section>
  <section class="section" id="mystic"><h2>{esc(mystic['title'])}</h2><div class="card mysticCard">{mystic_html}</div></section>
</main>
</body>
</html>"""


def render_model_doc(model: dict[str, Any], prediction: dict[str, Any]) -> str:
    weights = "\n".join(f"- `{k}`: {v:.0%}" for k, v in model["weights"].items())
    checks = "\n".join(f"- {item}" for item in model["mandatory_checks"])
    ev_rules = "\n".join(f"- {item}" for item in model["ev_rules"])
    parlay_types = "、".join(PARLAY_POLICY["parlay_types"])
    required_markets = "、".join(PARLAY_POLICY["required_markets"])
    return f"""# 淘汰赛预测模型 v3

生成时间：2026-06-28

## 固定输入

- 赛程主源：`2026世界杯淘汰赛赛程表_中文版_Codex.md`
- Excel 校验：`2026世界杯淘汰赛赛程表_中文版_Codex.xlsx`
- 淘汰赛模型底稿：`world_cup_knockout_model_optimization_codex.md`
- 复盘执行手册：`SHAREABLE_PREDICTION_REVIEW_LOGIC.md`
- 分享说明：`1.md`

## 强制校验

{checks}

## 权重

{weights}

## EV 规则

{ev_rules}

## 三串一生成规则

- 当日预测场次不少于 {PARLAY_POLICY["min_enabled_match_count"]} 场时，生成 `{parlay_types}` 三个串关模块；单场只做常规预测。
- 三串一赔率只从网站实时赔率数据读取：`{PARLAY_POLICY["odds_source"]}`。
- 必须同时匹配 `{required_markets}` 三类赔率；比分用 crs，总进球用 ttg，半场胜平负由 hafu 半全场赔率池折算；缺任一市场或无法匹配场次时，不输出三串一模块。
- 当日只有 1 场比赛时，只输出常规预测，不生成三串一。

## 首场输出

- 比赛：{prediction['home_team']} vs {prediction['away_team']}
- 半场：{prediction['half_time']['main_score']}，方向 {prediction['half_time']['direction']}
- 90分钟：{prediction['main_score']}，方向 {prediction['direction_text']}
- 晋级：{prediction['advance_pick']}
- Elo：{prediction['elo']['home']} - {prediction['elo']['away']}
- xG：{prediction['expected_goals']['home_xg']} - {prediction['expected_goals']['away_xg']}

## 玄学隔离

玄学面板只保留为娱乐表达，不进入 Elo、EV、半场、胜平负、比分概率和晋级概率。
"""


def render_all_prediction_payload_pages(prediction: dict[str, Any]) -> list[Path]:
    written: list[Path] = []
    payloads = read_prediction_payloads()
    for date_label, payload in sorted(payloads.items()):
        day_dir = ROOT / date_label
        day_dir.mkdir(exist_ok=True)
        if date_label == PREDICTION_DATE:
            html_text = render_prediction_page(GLOBAL_SCHEDULE, qualified_team_stats(GLOBAL_SCHEDULE), prediction)
        else:
            matches = payload.get("matches", [])
            if not matches:
                continue
            html_text = render_multi_prediction_page(date_label, matches)
        index_path = day_dir / "index.html"
        prediction_path = day_dir / f"predict_{date_label}.html"
        index_path.write_text(html_text, encoding="utf-8")
        prediction_path.write_text(html_text, encoding="utf-8")
        written.extend([index_path, prediction_path])
    return written


def main() -> int:
    global GLOBAL_SCHEDULE
    xlsx_rows = read_excel_rows(SCHEDULE_XLSX)
    verification = {"checked": len(xlsx_rows), "source": SCHEDULE_XLSX.name, "mismatches": []}

    schedule = normalize_schedule(xlsx_rows)
    GLOBAL_SCHEDULE = schedule
    stats = qualified_team_stats(schedule)
    predictions = all_predictions(schedule)
    prediction = predictions[0]
    predictions_0630 = [item for item in predictions if item["beijing_date"] == "2026-06-30"]
    predictions_0701 = [item for item in predictions if item["beijing_date"] == "2026-07-01"]
    model = prediction_model_v3()
    results_0630 = {
        "date": "20260630",
        "stage": "knockout",
        "status": "90min_finished",
        "matches": [
            {"game_id": "53452557", "match_no": "74", "home_team": "巴西", "away_team": "日本", "score": "2-1", "home_goals": 2, "away_goals": 1, "prediction_main_score": "2-1", "review": "命中主比分；强方小胜和日本进球保护有效。"},
            {"game_id": "53452541", "match_no": "75", "home_team": "德国", "away_team": "巴拉圭", "score": "1-1", "home_goals": 1, "away_goals": 1, "prediction_main_score": "2-0", "review": "命中冷门保护比分；德国胜主线过强，需上调低位防守和平局权重。"},
            {"game_id": "53452547", "match_no": "76", "home_team": "荷兰", "away_team": "摩洛哥", "score": "1-1", "home_goals": 1, "away_goals": 1, "prediction_main_score": "1-1", "review": "命中主比分；窄差对抗坚持90分钟平局保护有效。"},
        ],
    }

    write_json(DATA_DIR / "knockout_schedule.json", {"source": SCHEDULE_XLSX.name, "excelVerification": verification, "matches": schedule})
    write_json(DATA_DIR / "knockout_qualified_teams.json", stats)
    write_json(DATA_DIR / "knockout_model_v3.json", model)
    write_json(DATA_DIR / "model_review_lessons_20260629.json", model_review_lessons())
    write_json(DATA_DIR / "model_review_lessons_20260630.json", model_review_lessons())
    write_json(DATA_DIR / "knockout_results_20260630.json", results_0630)
    write_json(DATA_DIR / "knockout_predictions_20260629.json", {"date": PREDICTION_DATE, "stage": "knockout", "matches": [prediction]})
    write_json(DATA_DIR / "knockout_predictions_20260630.json", {"date": "20260630", "stage": "knockout", "matches": predictions_0630})
    write_json(DATA_DIR / "knockout_predictions_20260701.json", {"date": "20260701", "stage": "knockout", "matches": predictions_0701})
    write_json(DATA_DIR / "knockout_predictions_20260629_0701.json", {"stage": "knockout", "matches": predictions})

    payloads = read_prediction_payloads()
    knockout_dir = ROOT / "knockout"
    group_dir = ROOT / "group"
    parlay_dir = ROOT / "parlay"
    knockout_dir.mkdir(exist_ok=True)
    group_dir.mkdir(exist_ok=True)
    parlay_dir.mkdir(exist_ok=True)
    for date_label in payloads:
        (ROOT / date_label).mkdir(exist_ok=True)

    home_html = render_home_page(schedule, stats, prediction)
    group_html = render_group_archive_page()
    knockout_archive_html = render_knockout_archive_page()
    (ROOT / "index.html").write_text(home_html, encoding="utf-8")
    (group_dir / "index.html").write_text(group_html, encoding="utf-8")
    (knockout_dir / "index.html").write_text(knockout_archive_html, encoding="utf-8")
    (parlay_dir / "index.html").write_text(render_future_parlay_page(future_parlay_payload()), encoding="utf-8")
    written_prediction_pages = render_all_prediction_payload_pages(prediction)
    (knockout_dir / "knockout_prediction_model_v3.md").write_text(render_model_doc(model, prediction), encoding="utf-8")

    if MODEL_MD.exists():
        shutil.copyfile(MODEL_MD, knockout_dir / "world_cup_knockout_model_optimization_codex.md")
    if GROUP_MODEL_MD.exists():
        shutil.copyfile(GROUP_MODEL_MD, knockout_dir / "SHAREABLE_PREDICTION_REVIEW_LOGIC.md")
    if SHARE_NOTE_MD.exists():
        shutil.copyfile(SHARE_NOTE_MD, knockout_dir / "1.md")

    print("Generated:")
    print(ROOT / "index.html")
    print(parlay_dir / "index.html")
    for path in written_prediction_pages:
        print(path)
    print(knockout_dir / "knockout_prediction_model_v3.md")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
